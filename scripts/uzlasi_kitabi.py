# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — UZLASI KITABI URETICI

NE YAPAR
--------
gecis2_uzlasi.csv'deki uyusmazliklari, uzmanlarin karar verebilmesi icin
gereken HER SEYI iceren tek bir calisma kitabina donusturur:

  * iddia metni (tam)
  * altin etiket ve her iki uzmanin karari + gerekcesi
  * ATIF COZUMU: iddianin gosterdigi birim (iddia metninden cozulur)
  * O BIRIMIN KORPUSTAN TAM METNI  <-- ikinci gecis kitaplarinda yoktu
  * Gerekcede baska bir birim adi geciyorsa ONUN DA tam metni
  * bos UZLASI ve GEREKCE alanlari

NEDEN KORPUS METNI SART
-----------------------
Uyusmazliklarin bir kismi TANIM sorusu degil OLGU sorusudur. Ornek: 364'te
INS_MUH cumlenin 15.3.1'de oldugunu, ISG_UZM 15.1'de oldugunu soyluyor.
Bu bakilarak cozulur. Ikinci gecis kitaplari kaynak maddeyi gosteriyordu
ama uzmanin isaret ettigi ALTERNATIF birimi gostermiyordu; bu yuzden
olgusal uyusmazlik tanim uyusmazligi gibi gorundu.

KONTROL MADDELERI
-----------------
tabaka=K olan satirlar EKILMIS kontrollerdir; dogru cevap tasarim geregi
bilinir. Bunlar uzlasiya GIRMEZ, ayri sayfada kontrol performansi olarak
raporlanir. Uzlasiya sokmak, denetimin duyarlilik olcusunu yok eder.

KULLANIM
--------
    python -u scripts\\uzlasi_kitabi.py
    python -u scripts\\uzlasi_kitabi.py --cikti data\\iddialar\\uzlasi_kitabi.xlsx
"""
import argparse
import collections
import csv
import json
import os
import re
import sys
import unicodedata

BURASI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BURASI)

for _akis in (sys.stdout, sys.stderr):
    try:
        _akis.reconfigure(encoding="utf-8", errors="replace")
    except Exception:                                       # noqa: BLE001
        pass

try:
    from atif_coz import atif
except Exception as exc:                                    # noqa: BLE001
    print(f"! atif_coz.py import edilemedi: {exc}")
    sys.exit(1)

# gerekce metninde gecen birim referanslarini yakala: 15.3.1, 2.5.2.1(b), 45. madde
BIRIM_REF = re.compile(r"\b(\d+(?:\.\d+){1,4})\b|\b(\d{1,3})\s*(?:\.|inci|ıncı|nci|ncı|uncu|üncü)?\s*madde")

BASLIKLAR = [
    ("kod", 10), ("olgu_turu", 18), ("atif_var_mi", 12),
    ("iddia", 62), ("altin", 9),
    ("INS_MUH", 10), ("INS_MUH_gerekce", 46),
    ("ISG_UZM", 10), ("ISG_UZM_gerekce", 46),
    ("atif_cozumu", 14), ("atifin_tam_metni", 70),
    ("GERCEKTE_GECTIGI_BIRIMLER", 26), ("gectigi_birimin_tam_metni", 70),
    ("UZLASI", 12), ("UZLASI_GEREKCE", 46),
]

# ------------------------------------------------------- KONUM TESPITI
# Gerekce metninden birim adi tahmin etmek kirilgan: 364'te uzmanin
# isaret ettigi "15.3.1" korpusta o adla yok (bentler() farkli
# numaralandirmis). Dogru yontem, iddianin one surdugu cumleyi KORPUSUN
# TAMAMINDA aramak. Bu, uzmanin nesir gerekcesini ayristirmaya hic
# bagli degil ve olgusal uyusmazligi kesin olarak cozer.
SERH = re.compile(r"\((?:Değişik|Ek|Mülga|Yeniden düzenleme)[^)]{0,140}\)")
GORE = re.compile(r"\sgöre\s")


def _anahtar(t):
    t = unicodedata.normalize("NFC", t)
    for a, b in (("\u2019", "'"), ("\u2018", "'"), ("\u201c", '"'), ("\u201d", '"')):
        t = t.replace(a, b)
    t = SERH.sub(" ", t).lower().replace("\u0307", "")
    return re.sub(r"\s+", "", t)


def _icerik(iddia):
    m = GORE.search(iddia)
    return (iddia[m.end():] if m else iddia).strip().rstrip(".").strip()


def gectigi_birimler(iddia, kanun, korpus_a):
    """Iddianin one surdugu cumle, o kanunun HANGI birimlerinde geciyor."""
    ic = _anahtar(_icerik(iddia))
    if not ic:
        return []
    return [no for (k, no), a in korpus_a.items() if k == kanun and ic in a]



OLGU_BASLIK = [
    ("id", 8), ("uretim_sablonu", 20), ("iddia", 62),
    ("mevcut_altin", 12), ("kayitli_bent", 12), ("kayitli_bendin_metni", 66),
    ("GERCEK_BENT", 12), ("gercek_bendin_metni", 66),
    ("ONERILEN_ALTIN", 14), ("TEYIT", 12), ("TEYIT_GEREKCE", 40),
]


def olgu_turu(gerekce):
    g = gerekce.lower()
    if "değil" in g and ("yer alıyor" in g or "başlığı" in g or "devamında" in g):
        return "OLGUSAL (konum)"
    if any(w in g for w in ("kapsamında", "kapsamındaki", "düzeyi", "için geçerli",
                            "koşulu kaldır", "genellemiş", "genel kural")):
        return "kapsam kosulu"
    if any(w in g for w in ("gönderme", "önceki cümle", "önceki fıkra",
                            "tanımlanan", "bağlı")):
        return "anafora"
    return "siniflandirilamadi"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--uzlasi", default="sonuclar/gecis2_uzlasi.csv")
    ap.add_argument("--claims", default="data/iddialar/uretilen_iddialar_v6_onarilmis.csv")
    ap.add_argument("--korpus", default="data/korpus/korpus.jsonl")
    ap.add_argument("--cikti", default="data/iddialar/uzlasi_kitabi.xlsx")
    ap.add_argument("--adaylar", default="sonuclar/altin_adaylari.csv")
    ap.add_argument("--pdf-dir", default="data/kaynak_pdf")
    a = ap.parse_args()

    with open(a.uzlasi, encoding="utf-8-sig") as fh:
        satirlar = list(csv.DictReader(fh))
    with open(a.claims, encoding="utf-8-sig") as fh:
        C = {r["id"]: r for r in csv.DictReader(fh)}
    korpus = {}
    with open(a.korpus, encoding="utf-8-sig") as fh:
        for s in fh:
            if s.strip():
                v = json.loads(s)
                korpus[(v["kanun"], v["birim"])] = v

    korpus_a = {k: _anahtar(v["metin"]) for k, v in korpus.items()}
    print(f"uzlasi satiri: {len(satirlar)}   korpus: {len(korpus)} birim")

    olgusal_kodlar = set()
    if os.path.exists(a.adaylar):
        with open(a.adaylar, encoding="utf-8-sig") as fh:
            olgusal_kodlar = {r["id"] for r in csv.DictReader(fh)
                              if r.get("ETIKET_SONUCU") == "ALTIN DOGRU->YANLIS"}
        print(f"EK-5 olgusal maddeler (TANIM sayfasindan cikarildi): "
              f"{sorted(olgusal_kodlar, key=int)}")

    uzlasilik, kontroller, eksik = [], [], []
    for x in satirlar:
        if x["tabaka"] == "K":
            kontroller.append(x)
            continue
        if x["INS_MUH_doldurulmus_karar"] == x["ISG_UZM_doldurulmus__karar"]:
            continue                       # uyusmazlik yok
        if x["kod"] in olgusal_kodlar:
            continue        # EK-5 ile OLGUSAL olarak cozuldu -> OLGU sayfasinda
        kod = x["kod"]
        iddia_r = C.get(kod)
        if iddia_r is None:
            eksik.append(kod)
            continue
        iddia = iddia_r["iddia"]
        kanun, madde = atif(iddia)
        birim = korpus.get((kanun, madde)) if madde else None

        # KESIN KONUM: cumleyi korpusun tamaminda ara
        gecen = gectigi_birimler(iddia, kanun, korpus_a)
        baska = [g for g in gecen if g != madde]
        alt = korpus.get((kanun, baska[0])) if baska else None
        if not gecen:
            konum = "HICBIR BIRIMDE BULUNAMADI"
        elif madde and madde in gecen and not baska:
            konum = f"yalniz atif yapilan birimde ({kanun}/{madde})"
        elif madde and madde in gecen:
            konum = f"{kanun}/{madde} VE AYRICA: " + ", ".join(baska[:4])
        elif madde:
            konum = f"ATIF YANLIS -> gercekte: " + ", ".join(gecen[:4])
        else:
            konum = "atifsiz; gectigi birimler: " + ", ".join(gecen[:4])

        uzlasilik.append({
            "kod": kod,
            "olgu_turu": olgu_turu(x["INS_MUH_doldurulmus_gerekce"]),
            "atif_var_mi": "ATIF VAR" if madde else "ATIFSIZ",
            "iddia": iddia,
            "altin": x["gold"],
            "INS_MUH": x["INS_MUH_doldurulmus_karar"],
            "INS_MUH_gerekce": x["INS_MUH_doldurulmus_gerekce"],
            "ISG_UZM": x["ISG_UZM_doldurulmus__karar"],
            "ISG_UZM_gerekce": x["ISG_UZM_doldurulmus__gerekce"],
            "atif_cozumu": f"{kanun}/{madde}" if madde else f"{kanun} (belge)",
            "atifin_tam_metni": (birim or {}).get("metin", "[birim bulunamadi]"),
            "GERCEKTE_GECTIGI_BIRIMLER": konum,
            "gectigi_birimin_tam_metni": (alt or {}).get("metin", ""),
            "UZLASI": "", "UZLASI_GEREKCE": "",
        })

    print(f"\nuzlasiya girecek madde : {len(uzlasilik)}")
    print(f"kontrol (uzlasiya GIRMEZ): {len(kontroller)}  -> {[k['kod'] for k in kontroller]}")
    if eksik:
        print(f"! iddia kumesinde bulunamadi: {eksik}")

    print("\nOLGU TURU DAGILIMI (on siniflandirma — uzmanlar degistirebilir)")
    for t, n in collections.Counter(r["olgu_turu"] for r in uzlasilik).items():
        print(f"    {t:<22}{n}")
    print("\nATIF EKSENI")
    for t, n in collections.Counter(r["atif_var_mi"] for r in uzlasilik).items():
        print(f"    {t:<22}{n}")

    print("\nMADDE OZETI")
    for r in uzlasilik:
        print(f"    {r['kod']:<6}{r['olgu_turu']:<18}{r['atif_var_mi']:<10}"
              f"{r['atif_cozumu']:<14}{r['GERCEKTE_GECTIGI_BIRIMLER']}")

    os.makedirs(os.path.dirname(a.cikti) or ".", exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        yedek = os.path.splitext(a.cikti)[0] + ".csv"
        with open(yedek, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=[b for b, _ in BASLIKLAR])
            w.writeheader()
            w.writerows(uzlasilik)
        print(f"\n! openpyxl yok -> CSV yazildi: {yedek}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "TANIM"
    baslik_f = PatternFill("solid", fgColor="1F3864")
    doldur_f = PatternFill("solid", fgColor="FFF2CC")
    for c, (b, w) in enumerate(BASLIKLAR, 1):
        h = ws.cell(1, c, b)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = baslik_f
        h.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    for i, r in enumerate(uzlasilik, 2):
        for c, (b, _) in enumerate(BASLIKLAR, 1):
            cell = ws.cell(i, c, r[b])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if b in ("UZLASI", "UZLASI_GEREKCE"):
                cell.fill = doldur_f
        ws.row_dimensions[i].height = 150
    dv = DataValidation(type="list",
                        formula1='"DOGRU,YANLIS,CERCEVE_DISI"', allow_blank=True)
    ws.add_data_validation(dv)
    sut = get_column_letter([b for b, _ in BASLIKLAR].index("UZLASI") + 1)
    dv.add(f"{sut}2:{sut}{len(uzlasilik)+1}")
    ws.freeze_panes = "D2"

    ws2 = wb.create_sheet("KURAL")
    for i, satir in enumerate([
        "UZLASI KURAL NOTU — ONCE BU SAYFA, SONRA UZLASI SAYFASI",
        "",
        "Kural yazilmadan madde acilirsa kural maddelere gore secilmis olur",
        "ve denetim degerini kaybeder. Ayrintili secenekler icin:",
        "sonuclar/UZLASI_KURAL_NOTU.md",
        "",
        "Secilen secenek:",
        "Kuralin tam ifadesi:",
        "Gerekce (ilkeye dayali):",
        "Kapsam kosulu ile anafora ayni muameleyi gorecek mi:",
        "364 (olgusal) icin karar:",
        "",
        "Insaat muhendisi / tarih / imza:",
        "ISG uzmani / tarih / imza:",
    ], 1):
        ws2.cell(i, 1, satir)
    ws2.column_dimensions["A"].width = 70
    ws2.column_dimensions["B"].width = 80

    # ---------------------------------------------- OLGU SAYFASI (EK-5)
    olgu = []
    if os.path.exists(a.adaylar):
        try:
            import uret_iddia_v3_6 as UR
            from bent_bol import bent_bol
            t = UR.normalize(UR.pdf_metin(os.path.join(a.pdf_dir,
                                                       UR.LAWS["TBDY"]["dosya"])))
            RB = bent_bol(UR.bentler(t), t)
        except Exception as exc:                            # noqa: BLE001
            print(f"  ! rafine korpus kurulamadi ({exc}); metin sutunlari bos kalacak")
            RB = {}
        with open(a.adaylar, encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                if r.get("ETIKET_SONUCU") != "ALTIN DOGRU->YANLIS":
                    continue
                iddia_r = C.get(r["id"], {})
                olgu.append({
                    "id": r["id"],
                    "uretim_sablonu": r["uretim_sablonu"],
                    "iddia": iddia_r.get("iddia", r.get("iddia", "")),
                    "mevcut_altin": r["gold"],
                    "kayitli_bent": r["kayitli_bent"],
                    "kayitli_bendin_metni": RB.get(r["kayitli_bent"], {}).get("metin", ""),
                    "GERCEK_BENT": r["gercek_bent"],
                    "gercek_bendin_metni":
                        RB.get(r["gercek_bent"].split(",")[0].strip(), {}).get("metin", ""),
                    "ONERILEN_ALTIN": "YANLIS",
                    "TEYIT": "", "TEYIT_GEREKCE": "",
                })
    if olgu:
        wso = wb.create_sheet("OLGU")
        for c, (b, w) in enumerate(OLGU_BASLIK, 1):
            h = wso.cell(1, c, b)
            h.font = Font(bold=True, color="FFFFFF")
            h.fill = PatternFill("solid", fgColor="7F3300")
            h.alignment = Alignment(wrap_text=True, vertical="center")
            wso.column_dimensions[get_column_letter(c)].width = w
        for i, r in enumerate(olgu, 2):
            for c, (b, _) in enumerate(OLGU_BASLIK, 1):
                cell = wso.cell(i, c, r[b])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if b in ("TEYIT", "TEYIT_GEREKCE"):
                    cell.fill = doldur_f
            wso.row_dimensions[i].height = 150
        dvo = DataValidation(type="list",
                             formula1='"ONAY,RET,KARARSIZ"', allow_blank=True)
        wso.add_data_validation(dvo)
        so = get_column_letter([b for b, _ in OLGU_BASLIK].index("TEYIT") + 1)
        dvo.add(f"{so}2:{so}{len(olgu)+1}")
        wso.freeze_panes = "C2"

    ws3 = wb.create_sheet("KONTROLLER")
    ws3.append(["kod", "tabaka", "gold", "INS_MUH", "ISG_UZM", "not"])
    for k in kontroller:
        ws3.append([k["kod"], k["tabaka"], k["gold"],
                    k["INS_MUH_doldurulmus_karar"], k["ISG_UZM_doldurulmus__karar"],
                    "EKILMIS KONTROL — uzlasiya girmez, tespit performansi olarak raporlanir"])
    for c, w in zip("ABCDEF", (10, 9, 9, 11, 11, 80)):
        ws3.column_dimensions[c].width = w

    wb.save(a.cikti)
    print(f"\nyazildi: {a.cikti}")
    print(f"  TANIM       {len(uzlasilik)} madde — KURAL gerektirir")
    print(f"  OLGU        {len(olgu)} madde — kural gerektirmez, teyit yeter")
    print(f"  KONTROLLER  {len(kontroller)} madde — dokunulmaz")


if __name__ == "__main__":
    main()
