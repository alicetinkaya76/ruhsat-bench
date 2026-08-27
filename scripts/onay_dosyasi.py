# -*- coding: utf-8 -*-
"""
RUHSAT-Bench F3 — uzman onay dosyası üreteci.
Üretilen iddia CSV'sini, açılır menülü (ONAY/RET/DUZELT) bir Excel onay dosyasına çevirir.

Kullanım:
    python scripts/onay_dosyasi.py
    python scripts/onay_dosyasi.py --csv data/iddialar/uretilen_iddialar_v1.csv --out data/iddialar/onay_turu_v1.xlsx

Çift bağımsız kodlama için dosyayı iki kopya halinde iki uzmana verin (UZMAN sütununa ad yazılır);
dönen iki dosyadan kodlayıcılar-arası uyum (Cohen kappa) hesaplanır.
"""
import argparse, csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

AR = Font(name="Arial", size=10)
ARB = Font(name="Arial", size=10, bold=True)
HDR = PatternFill("solid", fgColor="D9E2F3")
GOLD_D = PatternFill("solid", fgColor="E2EFDA")
GOLD_Y = PatternFill("solid", fgColor="FCE4EC")
ONAY_F = PatternFill("solid", fgColor="FFF2CC")
thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default="data/iddialar/uretilen_iddialar_v1.csv")
    ap.add_argument("--out", default="data/iddialar/onay_turu_v1.xlsx")
    a = ap.parse_args()

    rows = list(csv.DictReader(open(a.csv, encoding="utf-8-sig")))
    wb = Workbook()

    ws = wb.active
    ws.title = "ONAY"
    basliklar = ["id", "kanun", "madde", "probe", "altın", "iddia", "kaynak alıntısı",
                 "değişiklik notu", "şablon", "ONAY", "DÜZELTME/GEREKÇE", "UZMAN"]
    for c, h in enumerate(basliklar, 1):
        cell = ws.cell(1, c, h)
        cell.font = ARB; cell.fill = HDR; cell.border = thin
    dv = DataValidation(type="list", formula1='"ONAY,RET,DUZELT,EMIN_DEGILIM"', allow_blank=True,
                        showDropDown=False, promptTitle="Karar",
                        prompt="ONAY: iddia ve etiket doğru | RET: kullanılmasın | DUZELT: küçük düzeltmeyle kullanılabilir")
    ws.add_data_validation(dv)
    for r, k in enumerate(rows, 2):
        degerler = [k["id"], k["kanun"], k["madde"], k["probe"], k["gold"], k["iddia"],
                    k["kaynak_alinti"], k.get("degisiklik_notu", ""), k.get("uretim_sablonu", ""),
                    "", "", ""]
        for c, v in enumerate(degerler, 1):
            cell = ws.cell(r, c, v)
            cell.font = AR; cell.border = thin
            if c in (6, 7):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 5).fill = GOLD_D if k["gold"] == "DOGRU" else GOLD_Y
        ws.cell(r, 10).fill = ONAY_F
        dv.add(ws.cell(r, 10))
    genislikler = [5, 9, 9, 13, 8, 62, 45, 14, 16, 14, 28, 10]
    for c, w in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:L{len(rows) + 1}"

    ws2 = wb.create_sheet("BENIOKU")
    talimat = [
        ("RUHSAT-Bench F3 — Uzman Onay Turu", ""),
        ("", ""),
        ("Görev", "Her satırdaki iddianın (F sütunu) altın etiketiyle (E sütunu) tutarlı, dilbilgisel olarak tam ve tek başına anlaşılır olup olmadığını değerlendirin."),
        ("Dayanak", "G sütunundaki kaynak alıntısı, iddianın türetildiği resmî metin parçasıdır; şüphede kaynağın PDF'ine gidin (data/kaynak_pdf/)."),
        ("ONAY", "İddia + etiket doğru, ifade temiz → sete girer."),
        ("RET", "Yanlış etiket, anlamsız/kirli ifade, bağlamsız parça → sete girmez."),
        ("DUZELT", "Küçük müdahaleyle kurtulur → K sütununa düzeltilmiş ifadeyi/gerekçeyi yazın."),
        ("EMIN_DEGILIM", "İkinci uzmana/karar toplantısına kalsın."),
        ("Çift kodlama", "Bu dosyanın iki kopyası iki uzmanca BAĞIMSIZ doldurulur (L sütununa ad). Kopyalar birleştirilip Cohen kappa raporlanır; uyuşmayanlar birlikte karara bağlanır."),
        ("Süre tahmini", f"{len(rows)} satır × ~15 sn ≈ {len(rows)*15//60} dakika/uzman."),
    ]
    for r, (ad, txt) in enumerate(talimat, 1):
        ws2.cell(r, 1, ad).font = ARB if ad and not txt else AR
        ws2.cell(r, 2, txt).font = AR
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 120

    os.makedirs(os.path.dirname(a.out), exist_ok=True)
    wb.save(a.out)
    print(f"yazıldı: {a.out}  ({len(rows)} satır)")


if __name__ == "__main__":
    main()
