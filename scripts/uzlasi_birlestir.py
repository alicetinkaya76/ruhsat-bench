# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — UZLASI BIRLESTIRME ve v7 ALTIN URETIMI

GIRDI
-----
Iki kodlayicinin doldurdugu uzlasi kitaplari (TANIM + OLGU sayfalari) ve
imzali kural notlari.

OLCULEN SONUC
-------------
  OLGU  : 7/7 oybirligi, hepsi ONAY  -> 7 altin DOGRU->YANLIS
  TANIM : 3/5 uyum (115, 122, 213 -> DOGRU)
          2/5 ayrisma (315, 417): kodlayici 1 YANLIS, kodlayici 2 DOGRU

Ayrisma UYGULAMA hatasi degil; iki kodlayici FARKLI KURAL secti
(1 -> Secenek 3 "atif ayrimi", 2 -> Secenek 1 "metinsel varlik yeter")
ve her biri kendi kuralini tutarli uyguladi.

BIRINCIL KURAL SECIMI — GEREKCE (sonuctan bagimsiz)
---------------------------------------------------
Secenek 3, atifsiz bir iddiada kapsam kosulunun dusurulup dusurulmedigine
bakmayi gerektirir. Bu SEMANTIK bir yargidir; otomatiklestirilemez ve
yalnizca denetlenen 138 maddede yapilmistir. Kalan 335 maddede
yapilmadigi icin, Secenek 3 uygulanirsa DENETLENEN ve DENETLENMEYEN
maddeler farkli olcutle etiketlenmis olur. Bu, iki kuraldan herhangi
birini secmekten daha kotudur.

Secenek 1 ise butun kumeye tekduze uygulanabilir: atif olgusal olarak
yanlissa (OLGU sayfasi, mekanik tespit) etiket doner, aksi halde donmez.

Bu nedenle BIRINCIL = Secenek 1 (kodlayici 2).
Secenek 3, DUYARLILIK kolu olarak raporlanir (v7b).

Secim, hangi sonucu verdigine bakilarak degil, hangisinin kume genelinde
tutarli uygulanabildigine bakilarak yapilmistir.

CIKTI
-----
  v7a (BIRINCIL)  : OLGU'nun 7 donusu             -> 223 DOGRU / 250 YANLIS
  v7b (DUYARLILIK): + 315 ve 417                  -> 221 DOGRU / 252 YANLIS

KULLANIM
--------
    python -u scripts\\uzlasi_birlestir.py --kitaplar k1.xlsx,k2.xlsx
"""
import argparse
import collections
import csv
import os
import sys

for _akis in (sys.stdout, sys.stderr):
    try:
        _akis.reconfigure(encoding="utf-8", errors="replace")
    except Exception:                                       # noqa: BLE001
        pass


def sayfa(wb, ad, anahtar_sut):
    ws = wb[ad]
    bas = [c.value for c in ws[1]]
    out = {}
    for r in range(2, ws.max_row + 1):
        sat = {b: (ws.cell(r, i + 1).value or "") for i, b in enumerate(bas)}
        k = str(sat.get(anahtar_sut, "")).strip()
        if k:
            out[k] = sat
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--kitaplar", required=True, help="virgulle ayrilmis iki xlsx")
    ap.add_argument("--claims", default="data/iddialar/uretilen_iddialar_v6_onarilmis.csv")
    ap.add_argument("--out-a", default="data/iddialar/uretilen_iddialar_v7a.csv")
    ap.add_argument("--out-b", default="data/iddialar/uretilen_iddialar_v7b.csv")
    ap.add_argument("--rapor", default="sonuclar/uzlasi_nihai.txt")
    a = ap.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("! openpyxl gerekli")
        sys.exit(1)

    yollar = [y.strip() for y in a.kitaplar.split(",")]
    if len(yollar) != 2:
        print("! tam iki kitap verin")
        sys.exit(1)
    W = [load_workbook(y) for y in yollar]

    R = []
    def y(s=""):
        print(s)
        R.append(s)

    y("=" * 76)
    y("UZLASI BIRLESTIRME")
    y("=" * 76)
    for i, p in enumerate(yollar, 1):
        y(f"  kodlayici {i}: {os.path.basename(p)}")

    # ------------------------------------------------------------- TANIM
    T = [sayfa(w, "TANIM", "kod") for w in W]
    ortak_t = sorted(set(T[0]) & set(T[1]), key=int)
    y(f"\nTANIM SAYFASI  (n={len(ortak_t)})")
    y(f"  {'kod':<7}{'kodlayici1':<12}{'kodlayici2':<12}{'durum'}")
    tanim_karar, ayrisan = {}, []
    for k in ortak_t:
        d1 = str(T[0][k].get("UZLASI", "")).strip().upper()
        d2 = str(T[1][k].get("UZLASI", "")).strip().upper()
        if d1 == d2:
            tanim_karar[k] = d1
            y(f"  {k:<7}{d1:<12}{d2:<12}uyum")
        else:
            ayrisan.append((k, d1, d2))
            y(f"  {k:<7}{d1:<12}{d2:<12}*** AYRISMA")
    y(f"  uyum: {len(tanim_karar)}/{len(ortak_t)}")

    # -------------------------------------------------------------- OLGU
    O = [sayfa(w, "OLGU", "id") for w in W]
    ortak_o = sorted(set(O[0]) & set(O[1]), key=int)
    onay, olgu_ayri = [], []
    for k in ortak_o:
        t1 = str(O[0][k].get("TEYIT", "")).strip().upper()
        t2 = str(O[1][k].get("TEYIT", "")).strip().upper()
        if t1 == t2 == "ONAY":
            onay.append(k)
        elif t1 != t2:
            olgu_ayri.append((k, t1, t2))
    y(f"\nOLGU SAYFASI  (n={len(ortak_o)})")
    y(f"  iki kodlayici da ONAY: {len(onay)}  -> {onay}")
    y(f"  ayrisma: {len(olgu_ayri)}")
    for k, t1, t2 in olgu_ayri:
        y(f"    {k}: {t1} / {t2}")

    # -------------------------------------------------------- v7 URETIMI
    with open(a.claims, encoding="utf-8-sig") as fh:
        C = list(csv.DictReader(fh))
    alan = list(C[0].keys()) + ["v7_not"]

    def uret(yol, ek_donen):
        n = 0
        satirlar = []
        for x in C:
            r = dict(x, v7_not="")
            if x["id"] in onay:
                r["gold"] = "YANLIS"
                r["v7_not"] = "EK-5 olgusal duzeltme (yanlis bent atfi), 2/2 ONAY"
                n += 1
            elif x["id"] in ek_donen:
                r["gold"] = "YANLIS"
                r["v7_not"] = "TANIM ayrismasi, Secenek 3 uygulanmis (duyarlilik kolu)"
                n += 1
            elif x["id"] in tanim_karar:
                r["v7_not"] = f"TANIM uzlasisi: {tanim_karar[x['id']]} (2/2 uyum)"
            satirlar.append(r)
        os.makedirs(os.path.dirname(yol) or ".", exist_ok=True)
        with open(yol, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=alan)
            w.writeheader()
            w.writerows(satirlar)
        d = collections.Counter(r["gold"] for r in satirlar)
        return n, d

    ay_kod = [k for k, _, _ in ayrisan]
    na, da = uret(a.out_a, set())
    nb, db = uret(a.out_b, set(ay_kod))

    y("\n" + "=" * 76)
    y("v7 ALTIN KUMELERI")
    y("=" * 76)
    y(f"  v6  (orijinal)     : 230 DOGRU / 243 YANLIS")
    y(f"  v7a (BIRINCIL)     : {da['DOGRU']} DOGRU / {da['YANLIS']} YANLIS"
      f"   ({na} donus: OLGU oybirligi)")
    y(f"  v7b (DUYARLILIK)   : {db['DOGRU']} DOGRU / {db['YANLIS']} YANLIS"
      f"   ({nb} donus: + {ay_kod})")
    y(f"\n  BIRINCIL KURAL: Secenek 1 (kodlayici 2).")
    y(f"  Gerekce: Secenek 3, atifsiz iddialarda kapsam dusurulmesine dair")
    y(f"  SEMANTIK yargi gerektirir; yalnizca denetlenen 138 maddede yapildi.")
    y(f"  Kalan 335 maddede yapilmadigi icin uygulanirsa denetlenen ve")
    y(f"  denetlenmeyen maddeler FARKLI OLCUTLE etiketlenmis olur.")
    y(f"  Secenek 1 butun kumeye tekduze uygulanabilir. Secim sonuca degil,")
    y(f"  tekduze uygulanabilirlige dayanir.")
    y(f"\n  yazildi: {a.out_a}")
    y(f"  yazildi: {a.out_b}")

    os.makedirs(os.path.dirname(a.rapor) or ".", exist_ok=True)
    with open(a.rapor, "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(R) + "\n")
    print(f"  yazildi: {a.rapor}")


if __name__ == "__main__":
    main()
