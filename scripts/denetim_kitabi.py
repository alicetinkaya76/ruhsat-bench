# -*- coding: utf-8 -*-
"""
RUHSAT-Bench F3 — KOR UZMAN DENETIM KITABI (iki bagimsiz kodlayici).

ornek_denetim.py --mod uzman'dan farklari
-----------------------------------------
1. ALTIN ETIKET GOSTERILMEZ. Eski kitapta "altin etiket" sutunu uzmanin
   onundeydi; bu bir capa (anchoring) yaratir ve iki uzman ayni capayi
   gordugu icin Cohen kappa yapay olarak sisar. Burada uzman kendi
   etiketini bagimsiz uretir, altinla karsilastirma SONRADAN yapilir.
2. IKI EKSEN. Uzman iki ayri yargida bulunur:
     KARAR  = iddia dogru mu?        {DOGRU, YANLIS, EMIN_DEGILIM}
     KALITE = madde saglam mi?       {TEMIZ, BAGLAMSIZ, KIRLI}
   Ikincisi dogruluktan bagimsizdir ve F3'un "kalan risk anlamsal duzeyde"
   iddiasini olculebilir kilar.
3. TASARIM AGIRLIKLARI YAZILIR. A tabakasi proba gore tabakali cekildigi
   icin basit ortalama YANLIdir; anahtar dosyasina her satirin N_h/n_h
   agirligi yazilir, kappa_birlestir.py agirlikli tahmin uretir.
4. HER UZMANA FARKLI SIRA. Ayni sira paylasilirsa yorgunluk/sira etkisi iki
   kodlayicida ESLESIR ve uyumu yapay olarak yukseltir. Sira uzman basina
   ayri karistirilir; birlestirme kod uzerinden yapilir.
5. OPSIYONEL DIKKAT TUZAGI (--tuzak N). Sayisi degistirilerek kendi
   alintisiyla celisir hale getirilmis N madde eklenir; dogru etiketi
   YANLIS'tir ve anahtarda isaretlidir. Uzman katmaninin POZITIF
   KONTROLUDUR: uzman bunlari yakalamiyorsa denetimin gucu bilinmiyordur.
   Tuzak satirlari hata orani tahminlerinden dislanir.

Gosterilen ama sizdirmayan alanlar: kanun, alintinin alindigi madde,
iddia, kaynak alintisi, (P6 icin) degisiklik notu.
GOSTERILMEYEN: gold, probe, uretim_sablonu, tabaka.

Kullanim:
    python scripts/denetim_kitabi.py --csv data/iddialar/uretilen_iddialar_v4_temiz.csv
    python scripts/denetim_kitabi.py --n-rastgele 100 --n-bayrak 50 --tuzak 8
"""
import argparse
import csv
import os
import random
import re
from collections import defaultdict

KARAR_SECENEK = "DOGRU,YANLIS,EMIN_DEGILIM"
KALITE_SECENEK = "TEMIZ,BAGLAMSIZ,KIRLI"


def kisalt(s, n):
    s = " ".join((s or "").split())
    return s if len(s) <= n else s[: n - 1] + "…"


def tuzak_uret(kayit, rnd):
    """Iddiadaki bir sayiyi degistirir; madde kendi alintisiyla celisir hale gelir."""
    iddia = " ".join((kayit.get("iddia") or "").split())
    adaylar = [m for m in re.finditer(r"\b\d{1,4}\b", iddia)]
    if not adaylar:
        return None
    m = rnd.choice(adaylar)
    eski = int(m.group(0))
    yeni = eski * 2 if eski * 2 <= 9999 else max(1, eski // 2)
    if yeni == eski:
        yeni = eski + 3
    return iddia[: m.start()] + str(yeni) + iddia[m.end():]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default="data/iddialar/uretilen_iddialar_v4_temiz.csv")
    ap.add_argument("--bayrak", default="sonuclar/konsensus_bayrak.csv")
    ap.add_argument("--n-rastgele", type=int, default=100)
    ap.add_argument("--n-bayrak", type=int, default=50)
    ap.add_argument("--tuzak", type=int, default=0)
    ap.add_argument("--uzman", default="UZMAN_1,UZMAN_2")
    ap.add_argument("--tohum", type=int, default=20260728)
    ap.add_argument("--out-dir", default="data/iddialar")
    ap.add_argument("--anahtar", default="sonuclar/denetim_anahtar_v2.csv")
    ap.add_argument("--rapor", default="sonuclar/denetim_kitabi_rapor.txt")
    a = ap.parse_args()

    with open(a.csv, encoding="utf-8-sig") as fh:
        satirlar = list(csv.DictReader(fh))
    kayit = {s["id"]: s for s in satirlar}

    bayrakli = set()
    if os.path.exists(a.bayrak):
        with open(a.bayrak, encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                if r["id"] in kayit:
                    bayrakli.add(r["id"])

    L = []

    def e(s=""):
        L.append(s)
        print(s)

    e("=" * 78)
    e("RUHSAT-Bench — KOR UZMAN DENETIM KITABI")
    e("=" * 78)
    e(f"kaynak kume: {a.csv}  ({len(satirlar)} iddia)")
    e(f"bayrakli (temiz kumede kalan): {len(bayrakli)}")

    rnd = random.Random(a.tohum)

    # ---- tabaka x probe hucreleri --------------------------------------
    hucre = defaultdict(list)
    for s in satirlar:
        t = "B_bayrakli" if s["id"] in bayrakli else "A_rastgele"
        hucre[(t, s.get("probe", "?"))].append(s["id"])
    for v in hucre.values():
        v.sort(key=int)

    def cek(tabaka, toplam):
        gr = {k: v for k, v in hucre.items() if k[0] == tabaka}
        if not gr or toplam <= 0:
            return {}
        pay = max(1, toplam // len(gr))
        secim = {}
        for k, v in sorted(gr.items()):
            n = min(pay, len(v))
            for cid in rnd.sample(v, n):
                secim[cid] = k
        artan = [(cid, k) for k, v in sorted(gr.items()) for cid in v if cid not in secim]
        rnd.shuffle(artan)
        for cid, k in artan[: max(0, toplam - len(secim))]:
            secim[cid] = k
        return secim

    secim = {}
    secim.update(cek("A_rastgele", a.n_rastgele))
    secim.update(cek("B_bayrakli", a.n_bayrak))

    n_hucre = defaultdict(int)
    for k in secim.values():
        n_hucre[k] += 1

    e()
    e("[1] TABAKA x PROBE  (N = evrende, n = ornekte, agirlik = N/n)")
    e(f"    {'tabaka':<12} {'probe':<16} {'N':>5} {'n':>4} {'agirlik':>8}")
    for k in sorted(hucre):
        N, n = len(hucre[k]), n_hucre.get(k, 0)
        ag = (N / n) if n else float("nan")
        e(f"    {k[0]:<12} {k[1]:<16} {N:>5} {n:>4} {ag:>8.2f}")
    e(f"  toplam secilen: {len(secim)}  "
      f"(A={sum(1 for v in secim.values() if v[0]=='A_rastgele')}, "
      f"B={sum(1 for v in secim.values() if v[0]=='B_bayrakli')})")

    # ---- tuzaklar ------------------------------------------------------
    tuzaklar = {}
    if a.tuzak > 0:
        havuz = [s for s in satirlar
                 if s.get("gold") == "DOGRU" and s["id"] not in secim
                 and re.search(r"\b\d{1,4}\b", s.get("iddia") or "")]
        rnd.shuffle(havuz)
        for s in havuz:
            if len(tuzaklar) >= a.tuzak:
                break
            yeni = tuzak_uret(s, rnd)
            if yeni:
                kod = f"T{len(tuzaklar)+1}"
                tuzaklar[kod] = dict(s, id=kod, iddia=yeni, gold="YANLIS")
        e()
        e(f"[2] DIKKAT TUZAGI: {len(tuzaklar)} madde eklendi (dogru etiket = YANLIS).")
        e("  Bunlar hata orani tahminlerinden dislanir; uzman katmaninin")
        e("  pozitif kontrolu olarak kullanilir.")

    # ---- anahtar -------------------------------------------------------
    os.makedirs(os.path.dirname(a.anahtar) or ".", exist_ok=True)
    with open(a.anahtar, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["kod", "tabaka", "probe", "probe_alt", "kanun", "madde",
                    "gold", "N_hucre", "n_hucre", "agirlik", "tuzak"])
        for cid, k in sorted(secim.items(), key=lambda x: int(x[0])):
            s = kayit[cid]
            N, n = len(hucre[k]), n_hucre[k]
            w.writerow([cid, k[0], s.get("probe", ""), s.get("probe_alt", ""),
                        s.get("kanun", ""), s.get("madde", ""), s.get("gold", ""),
                        N, n, f"{N/n:.6f}", 0])
        for kod, s in tuzaklar.items():
            w.writerow([kod, "T_tuzak", s.get("probe", ""), s.get("probe_alt", ""),
                        s.get("kanun", ""), s.get("madde", ""), "YANLIS", 0, 0, "0", 1])
    e()
    e(f"anahtar: {a.anahtar}  <-- UZMANLARA VERILMEZ")

    # ---- kitaplar ------------------------------------------------------
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    AR = Font(name="Arial", size=10)
    ARB = Font(name="Arial", size=10, bold=True)
    thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    tum = {cid: kayit[cid] for cid in secim}
    tum.update(tuzaklar)
    uzmanlar = [u.strip() for u in a.uzman.split(",") if u.strip()]
    os.makedirs(a.out_dir, exist_ok=True)
    yollar = []

    for ui, uzman in enumerate(uzmanlar):
        sira = list(tum)
        random.Random(a.tohum + 1000 * (ui + 1)).shuffle(sira)

        wb = Workbook()
        ws = wb.active
        ws.title = "DENETIM"
        basliklar = ["sira", "kod", "belge", "alıntının maddesi", "İDDİA",
                     "KAYNAK ALINTISI", "ek bilgi", "KARAR", "KALİTE",
                     "GEREKÇE / DÜZELTME", "UZMAN"]
        for c, h in enumerate(basliklar, 1):
            cell = ws.cell(1, c, h)
            cell.font = ARB
            cell.fill = PatternFill("solid", fgColor="D9E2F3")
            cell.border = thin
        dv_k = DataValidation(type="list", formula1=f'"{KARAR_SECENEK}"', allow_blank=True,
                              promptTitle="Karar",
                              prompt="Iddia mevzuata gore dogru mu? Emin degilseniz EMIN_DEGILIM.")
        dv_q = DataValidation(type="list", formula1=f'"{KALITE_SECENEK}"', allow_blank=True,
                              promptTitle="Kalite",
                              prompt="TEMIZ: tek basina anlasilir tam hukum. BAGLAMSIZ: baglami olmadan anlasilmiyor. KIRLI: metin artigi.")
        ws.add_data_validation(dv_k)
        ws.add_data_validation(dv_q)

        for i, kod in enumerate(sira, 1):
            s = tum[kod]
            ek = s.get("degisiklik_notu", "") or ""
            degerler = [i, kod, s.get("kanun", ""), s.get("madde", ""),
                        " ".join((s.get("iddia") or "").split()),
                        " ".join((s.get("kaynak_alinti") or "").split()),
                        ek, "", "", "", uzman]
            for c, v in enumerate(degerler, 1):
                cell = ws.cell(i + 1, c, v)
                cell.font = AR
                cell.border = thin
                if c in (5, 6):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(i + 1, 8).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(i + 1, 9).fill = PatternFill("solid", fgColor="FFF2CC")
            dv_k.add(ws.cell(i + 1, 8))
            dv_q.add(ws.cell(i + 1, 9))
        for c, w_ in enumerate([6, 7, 9, 14, 62, 48, 16, 15, 13, 32, 11], 1):
            ws.column_dimensions[get_column_letter(c)].width = w_
        ws.freeze_panes = "E2"
        ws.auto_filter.ref = f"A1:K{len(sira)+1}"

        ws2 = wb.create_sheet("BENIOKU")
        dk = len(sira) * 35 // 60
        talimat = [
            ("RUHSAT-Bench — Kör Uzman Denetimi", ""),
            ("", ""),
            ("Amaç", "Türkçe mühendislik mevzuatı üzerine kurulmuş bir değerlendirme kümesinin "
                     "etiket geçerliliğini ölçmek. Sizin kararınız referans kabul edilecek."),
            ("Bağımsızlık", "Bu dosyayı tek başınıza doldurun. Diğer uzmanla dosya dolana kadar "
                            "madde üzerine konuşmayın; uyuşmazlıklar sonradan birlikte karara bağlanacak."),
            ("", ""),
            ("KARAR sütunu", "İDDİA (E sütunu) mevzuata göre doğru mu?"),
            ("  DOGRU", "İddia, atıf yaptığı belge/madde açısından doğrudur."),
            ("  YANLIS", "İddia yanlıştır: sayı, yıl, atıf yapılan belge veya madde numarası hatalı, "
                         "ya da böyle bir hüküm/belge yoktur."),
            ("  EMIN_DEGILIM", "Verilen bilgiyle karara varamıyorsanız. Tahmin etmeyin."),
            ("", ""),
            ("KALİTE sütunu", "İddianın DOĞRULUĞUNDAN bağımsız olarak, madde bir ölçüm maddesi olarak sağlam mı?"),
            ("  TEMIZ", "Tek başına anlaşılır, tam bir hüküm cümlesi."),
            ("  BAGLAMSIZ", "Kendi başına anlaşılmıyor: eksik özne, 'bu durumda' gibi bağlam gerektiren başlangıç."),
            ("  KIRLI", "Metin artığı: başlık parçası, iki fıkranın birbirine karışması, "
                        "bölünmüş kelime, OCR bozukluğu."),
            ("", ""),
            ("Dayanak", "F sütunu, iddianın türetildiği resmî metin parçasıdır. D sütunu o parçanın "
                        "alındığı maddedir. İddianın kendi içinde başka bir maddeye/belgeye atıf yapması "
                        "mümkündür; bunun doğru olup olmadığı sizin değerlendirmenize kalmıştır."),
            ("Şüphede", "Kaynak belgeye gidin. Karar veremiyorsanız EMIN_DEGILIM işaretleyip "
                        "GEREKÇE sütununa nedenini yazın."),
            ("", ""),
            ("Süre", f"{len(sira)} satır × ~35 sn ≈ {dk} dakika."),
            ("Not", "Altın etiket bilerek gösterilmemiştir; kararınızın bağımsız olması ölçümün kendisidir."),
        ]
        for r, (ad, txt) in enumerate(talimat, 1):
            ws2.cell(r, 1, ad).font = ARB if ad and not txt else AR
            ws2.cell(r, 2, txt).font = AR
            ws2.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 110

        yol = os.path.join(a.out_dir, f"denetim_{uzman}.xlsx")
        wb.save(yol)
        yollar.append(yol)
        e(f"yazildi: {yol}  ({len(sira)} satir, sira uzmana ozel)")

    os.makedirs(os.path.dirname(a.rapor) or ".", exist_ok=True)
    with open(a.rapor, "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(L) + "\n")
    print(f"yazildi: {a.rapor}")


if __name__ == "__main__":
    main()
