# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — yanlilik duzeltmeli model-konsensus on-taramasi (F3 destek araci, v2).

v1'in sorunu: her cevaplayan modelin oyu esit sayiliyordu. Oysa yerel modellerin bir
kismi dejenere davraniyor (her seye YANLIS diyen llama3.2:3b gibi) ya da tek yonlu
yanli. Bu oylar altin etiket hakkinda bilgi tasimaz; dahasi hep-YANLIS diyen bir model,
altini DOGRU olan her maddede otomatik "karsi" oy uretir. Sonuc: bayraklar altin
hatasini degil, model yanliligini olcer.

v2 sirasiyla:
  1. Her model x kosul hucresi icin tanilayici olcer: kapsam, baskin etiket payi,
     sinif-kosullu dogruluk (acc_D, acc_Y), dengeli dogruluk, Youden J = 2*bacc-1.
  2. Bilgisiz hucreleri eler (dusuk kapsam / dejenere baskin etiket / J esigi alti).
  3. Her modelden en bilgili kosulu secer (v1 kor sekilde E2'yi tercih ediyordu).
  4. Kalan oylari J agirligiyla toplar; agirlikli karsi-oy payi esigi asanlari isaretler.
  5. Bayraklari altin sinifina / probe'a / kanuna gore ayirir; asimetri kalirsa uyarir.
  6. Bayraklari oncelik sirasina dizer (uzman once en bilgili itirazlara baksin).

Cikti:
  sonuclar/konsensus_tani.csv            model x kosul tanilayici tablo (makale Tablosu)
  sonuclar/konsensus_bayrak.csv          bayrakli maddeler, oncelik sirali
  sonuclar/konsensus_rapor.txt           ozet rapor (sohbete yapistirmak icin)
  data/iddialar/onay_turu_v2_uyarili.xlsx  uzmanlara verilecek dosya

Kullanim:
    python scripts/konsensus_uyari_v2.py
    python scripts/konsensus_uyari_v2.py --esik 0.7 --dejenere 0.85 --min-j 0.10
"""
import argparse
import csv
import json
import os
from collections import defaultdict, Counter

ETIKETLER = ("DOGRU", "YANLIS")


def tani_hesapla(kararlar, gold, n_toplam):
    """kararlar: {id: 'DOGRU'|'YANLIS'} -> tanilayici sozluk."""
    n = len(kararlar)
    dagilim = Counter(kararlar.values())
    baskin_pay = (max(dagilim.values()) / n) if n else 0.0
    d_t = sum(1 for i in kararlar if gold.get(i) == "DOGRU")
    y_t = sum(1 for i in kararlar if gold.get(i) == "YANLIS")
    d_h = sum(1 for i, k in kararlar.items() if gold.get(i) == "DOGRU" and k == "DOGRU")
    y_h = sum(1 for i, k in kararlar.items() if gold.get(i) == "YANLIS" and k == "YANLIS")
    acc_d = (d_h / d_t) if d_t else None
    acc_y = (y_h / y_t) if y_t else None
    if acc_d is None or acc_y is None:
        bacc, j = None, None
    else:
        bacc = (acc_d + acc_y) / 2.0
        j = 2 * bacc - 1
    return {
        "n": n,
        "kapsam": n / n_toplam if n_toplam else 0.0,
        "baskin": dagilim.most_common(1)[0][0] if n else "-",
        "baskin_pay": baskin_pay,
        "dogru_pay": dagilim.get("DOGRU", 0) / n if n else 0.0,
        "acc_D": acc_d,
        "acc_Y": acc_y,
        "bacc": bacc,
        "J": j,
    }


def f(x, nd=3):
    return "-" if x is None else f"{x:.{nd}f}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--jsonl", default="sonuclar/konsensus.jsonl")
    ap.add_argument("--csv", default="data/iddialar/uretilen_iddialar_v1.csv")
    ap.add_argument("--onay", default="data/iddialar/onay_turu_v1.xlsx")
    ap.add_argument("--out", default="data/iddialar/onay_turu_v2_uyarili.xlsx")
    ap.add_argument("--tani", default="sonuclar/konsensus_tani.csv")
    ap.add_argument("--bayrak", default="sonuclar/konsensus_bayrak.csv")
    ap.add_argument("--rapor", default="sonuclar/konsensus_rapor.txt")
    ap.add_argument("--esik", type=float, default=0.7, help="agirlikli karsi-oy payi esigi")
    ap.add_argument("--dejenere", type=float, default=0.85, help="baskin etiket payi ust siniri")
    ap.add_argument("--min-kapsam", type=float, default=0.30, help="asgari cevaplama orani")
    ap.add_argument("--min-j", type=float, default=0.10, help="asgari Youden J")
    ap.add_argument("--min-oy", type=int, default=3, help="asgari bilgili oy sayisi")
    a = ap.parse_args()

    satirlar = []
    with open(a.csv, encoding="utf-8-sig") as fh:
        for k in csv.DictReader(fh):
            satirlar.append(k)
    gold = {k["id"]: k["gold"] for k in satirlar}
    meta = {k["id"]: k for k in satirlar}
    n_toplam = len(gold)

    hucre = defaultdict(dict)  # (model, kosul) -> {id: karar}
    ham_sayac = Counter()
    with open(a.jsonl, encoding="utf-8") as fh:
        for satir in fh:
            satir = satir.strip()
            if not satir:
                continue
            r = json.loads(satir)
            ham_sayac[(r.get("model"), r.get("kosul"))] += 1
            if r.get("karar") in ETIKETLER and str(r.get("id")) in gold:
                hucre[(r["model"], r["kosul"])][str(r["id"])] = r["karar"]

    taniler = {}
    for anahtar, kararlar in hucre.items():
        taniler[anahtar] = tani_hesapla(kararlar, gold, n_toplam)

    # --- eleme ---
    for anahtar, t in taniler.items():
        gerekce = []
        if t["kapsam"] < a.min_kapsam:
            gerekce.append(f"dusuk kapsam ({t['kapsam']:.0%})")
        if t["baskin_pay"] >= a.dejenere:
            gerekce.append(f"dejenere ({t['baskin']} %{t['baskin_pay']*100:.0f})")
        if t["J"] is None:
            gerekce.append("J hesaplanamadi")
        elif t["J"] < a.min_j:
            gerekce.append(f"bilgisiz (J={t['J']:.2f})")
        t["elendi"] = bool(gerekce)
        t["gerekce"] = "; ".join(gerekce) if gerekce else "gecerli"

    # --- model basina en bilgili kosul ---
    en_iyi = {}
    for (model, kosul), t in taniler.items():
        if t["elendi"]:
            continue
        onceki = en_iyi.get(model)
        if onceki is None or t["J"] > taniler[(model, onceki)]["J"]:
            en_iyi[model] = kosul
    secili = [(m, k) for m, k in en_iyi.items()]

    esnek = False
    min_oy = a.min_oy
    if len(secili) < a.min_oy:
        esnek = True
        min_oy = max(2, len(secili))

    # --- oylama ---
    bayrak = {}
    oy_dagilim = Counter()
    naif = {}  # v1 taklidi: tum cevaplayanlar esit, E2 tercihli
    for cid, g in gold.items():
        # v2 agirlikli
        toplam_w, karsi_w, oy_sayisi, karsi_modeller = 0.0, 0.0, 0, []
        for model, kosul in secili:
            karar = hucre[(model, kosul)].get(cid)
            if not karar:
                continue
            w = max(taniler[(model, kosul)]["J"], 0.01)
            toplam_w += w
            oy_sayisi += 1
            if karar != g:
                karsi_w += w
                karsi_modeller.append(f"{model}/{kosul}")
        oy_dagilim[oy_sayisi] += 1
        if oy_sayisi >= min_oy and toplam_w > 0:
            pay = karsi_w / toplam_w
            if pay >= a.esik:
                bayrak[cid] = {
                    "pay": pay,
                    "oy": oy_sayisi,
                    "agirlik": toplam_w,
                    "modeller": ", ".join(karsi_modeller),
                }
        # v1 taklidi
        n_kararlar = []
        for model in {m for m, _ in hucre}:
            k2 = hucre.get((model, "E2"), {}).get(cid)
            k1 = hucre.get((model, "E1"), {}).get(cid)
            karar = k2 or k1
            if karar:
                n_kararlar.append(karar)
        if len(n_kararlar) >= 3:
            k = sum(1 for x in n_kararlar if x != g)
            if k / len(n_kararlar) >= a.esik:
                naif[cid] = k / len(n_kararlar)

    # --- kirilimlar ---
    def kirilim(alan):
        toplam, isaretli = Counter(), Counter()
        for cid, k in meta.items():
            v = k.get(alan, "?") if alan != "gold" else gold[cid]
            toplam[v] += 1
            if cid in bayrak:
                isaretli[v] += 1
        return toplam, isaretli

    g_top, g_bay = kirilim("gold")
    p_top, p_bay = kirilim("probe")
    k_top, k_bay = kirilim("kanun")

    ng_top, ng_bay = Counter(), Counter()
    for cid in meta:
        ng_top[gold[cid]] += 1
        if cid in naif:
            ng_bay[gold[cid]] += 1

    # --- rapor ---
    L = []
    ekle = L.append
    ekle("=" * 78)
    ekle("RUHSAT-Bench — yanlilik duzeltmeli konsensus on-taramasi (v2)")
    ekle("=" * 78)
    ekle(f"iddia sayisi: {n_toplam} | model x kosul hucresi: {len(taniler)}")
    ekle("")
    ekle("[1] MODEL x KOSUL TANILAYICI")
    ekle(f"{'model':<34}{'ks':<4}{'n':>5}{'kaps':>7}{'baskin':>16}{'accD':>7}{'accY':>7}{'bacc':>7}{'J':>7}  durum")
    for (model, kosul), t in sorted(taniler.items(), key=lambda x: (x[0][0], x[0][1])):
        ekle(f"{model[:33]:<34}{kosul:<4}{t['n']:>5}{t['kapsam']*100:>6.0f}%"
             f"{t['baskin'] + ' %' + format(t['baskin_pay']*100, '.0f'):>16}"
             f"{f(t['acc_D'], 2):>7}{f(t['acc_Y'], 2):>7}{f(t['bacc'], 2):>7}{f(t['J'], 2):>7}  {t['gerekce']}")
    ekle("")
    ekle(f"secilen oy veren hucreler ({len(secili)}):")
    for model, kosul in sorted(secili):
        ekle(f"  {model} / {kosul}  (J={taniler[(model, kosul)]['J']:.2f})")
    if esnek:
        ekle(f"  ! UYARI: bilgili model sayisi {a.min_oy} esiginin altinda; min-oy {min_oy}'e dusuruldu.")
    ekle("")
    ekle("[2] BAYRAK SONUCU")
    ekle(f"  v2 (agirlikli, elemeli) : {len(bayrak)} madde  (%{len(bayrak)/n_toplam*100:.1f})")
    ekle(f"  v1 (naif, esit oy)      : {len(naif)} madde  (%{len(naif)/n_toplam*100:.1f})")
    ortak = set(bayrak) & set(naif)
    ekle(f"  ortak                   : {len(ortak)} | yalniz v2: {len(set(bayrak)-set(naif))} | yalniz v1: {len(set(naif)-set(bayrak))}")
    ekle("")
    ekle("[3] ALTIN SINIFINA GORE (yanlilik testi)")
    for g in ("DOGRU", "YANLIS"):
        o1 = (ng_bay[g] / ng_top[g] * 100) if ng_top[g] else 0
        o2 = (g_bay[g] / g_top[g] * 100) if g_top[g] else 0
        ekle(f"  gold={g:<6} n={g_top[g]:<5} v1 bayrak={ng_bay[g]:<4}(%{o1:.1f})   v2 bayrak={g_bay[g]:<4}(%{o2:.1f})")
    a1 = abs((ng_bay['DOGRU']/max(ng_top['DOGRU'],1)) - (ng_bay['YANLIS']/max(ng_top['YANLIS'],1))) * 100
    a2 = abs((g_bay['DOGRU']/max(g_top['DOGRU'],1)) - (g_bay['YANLIS']/max(g_top['YANLIS'],1))) * 100
    ekle(f"  sinif asimetrisi: v1 {a1:.1f} puan -> v2 {a2:.1f} puan")
    if a2 > 8:
        fazla = "YANLIS" if (g_bay["YANLIS"] / max(g_top["YANLIS"], 1)) > (g_bay["DOGRU"] / max(g_top["DOGRU"], 1)) else "DOGRU"
        ekle(f"  ! Duzeltmeden SONRA da asimetri var ve yonu: gold={fazla} maddeler daha cok bayrakli.")
        ekle(f"    Bu, elemeden sonra kalan oy verenlerin ARTIK yanliligidir; J agirligi")
        ekle(f"    bilgililigi duzeltir, yanliligi duzeltmez. Bayraklari altin hatasi")
        ekle(f"    kaniti saymayin.")
    ekle("")
    ekle("[4] PROBE KIRILIMI")
    for p, n in sorted(p_top.items(), key=lambda x: -x[1]):
        ekle(f"  {p:<16} n={n:<5} bayrak={p_bay[p]:<4} (%{p_bay[p]/n*100:.1f})")
    ekle("")
    ekle("[5] KAYNAK KIRILIMI")
    for k, n in sorted(k_top.items(), key=lambda x: -x[1]):
        ekle(f"  {k:<28} n={n:<5} bayrak={k_bay[k]:<4} (%{k_bay[k]/n*100:.1f})")
    ekle("")
    ekle("[6] EN ONCELIKLI 20 BAYRAK")
    sirali = sorted(bayrak.items(), key=lambda x: (-x[1]["pay"], -x[1]["agirlik"]))
    for cid, b in sirali[:20]:
        m = meta[cid]
        ekle(f"  #{cid:<5} {m.get('kanun','')[:18]:<19}{m.get('madde','')[:10]:<11}{m.get('probe','')[:12]:<13}"
             f"gold={gold[cid]:<7}karsi=%{b['pay']*100:.0f} ({b['oy']} oy)")
    if len(sirali) > 20:
        ekle(f"  ... (+{len(sirali)-20})")
    ekle("")
    ekle("[7] YORUM")
    ekle(f"  Bayrak = 'altin yanlis' demek DEGILDIR; en fazla 'once buna bak' demektir.")
    ekle(f"  Pilotta 60 el-dogrulanmis maddede gercek altin hatasi 1 taneydi (~%1.7).")
    ekle("")
    ekle("[8] KULLANILABILIRLIK KARARI")
    en_iyi_j = max([taniler[c]["J"] for c in secili], default=0.0)
    ekle(f"  secilen hucrelerin en yuksek Youden J degeri: {en_iyi_j:.2f}")
    if en_iyi_j < 0.20:
        ekle(f"  ! HICBIR OY VEREN YETERINCE BILGILI DEGIL (J<0.20 => sansa cok yakin).")
        ekle(f"    Zayif oy verenlerden olusan bir konsensus, altin hatasini rastgele")
        ekle(f"    secimden anlamli olcude iyi bulamaz. Bu bayrak listesini uzman")
        ekle(f"    triyajinin SIRASI olarak kullanmayin; dogrulama icin tabakali")
        ekle(f"    rastgele ornek kullanin (bkz. scripts/ornek_denetim.py).")
    # probe yogunlasmasi: bayrak orani probe'a gore cok degisiyorsa, bayraklar
    # altin hatasini degil probe zorlugunu izliyor olabilir.
    oranlar = {p: p_bay[p] / n for p, n in p_top.items() if n >= 20}
    if len(oranlar) >= 3:
        en_u, en_a = max(oranlar.values()), max(min(oranlar.values()), 1e-9)
        ust = max(oranlar, key=oranlar.get)
        alt = min(oranlar, key=oranlar.get)
        ekle(f"  probe yogunlasmasi: en yuksek {ust} %{en_u*100:.1f} / en dusuk {alt} %{en_a*100:.1f}"
             f"  (oran x{en_u/en_a:.1f})")
        if en_u / en_a >= 3:
            ekle(f"  ! Bayraklar probe turune gore x{en_u/en_a:.1f} degisiyor. Altin hatasinin probe")
            ekle(f"    turune bu kadar bagli olmasi beklenmez; buna karsilik modellerin")
            ekle(f"    KUCUK sozcuksel degisiklikle yanlislanan iddialari (referans/sayi/tarih")
            ekle(f"    degistirme) kacirmasi beklenir. Yani bu bayraklar buyuk olasilikla")
            ekle(f"    ALTIN HATASINI degil PROBE ZORLUGUNU olcuyor. Bayrakli maddeleri")
            ekle(f"    setten atmak, kiyaslamayi en ayirt edici maddelerinden yoksun")
            ekle(f"    birakir ve model basarimini yapay olarak yukseltir.")
    rapor = "\n".join(L)
    print(rapor)

    os.makedirs(os.path.dirname(a.rapor) or ".", exist_ok=True)
    with open(a.rapor, "w", encoding="utf-8") as fh:
        fh.write(rapor + "\n")

    # --- tani csv ---
    with open(a.tani, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["model", "kosul", "n_cevap", "kapsam", "baskin_etiket", "baskin_pay",
                    "dogru_pay", "acc_DOGRU", "acc_YANLIS", "dengeli_dogruluk", "youden_J",
                    "elendi", "gerekce", "oy_kullandi"])
        for (model, kosul), t in sorted(taniler.items()):
            w.writerow([model, kosul, t["n"], f"{t['kapsam']:.4f}", t["baskin"],
                        f"{t['baskin_pay']:.4f}", f"{t['dogru_pay']:.4f}",
                        f(t["acc_D"], 4), f(t["acc_Y"], 4), f(t["bacc"], 4), f(t["J"], 4),
                        "EVET" if t["elendi"] else "HAYIR", t["gerekce"],
                        "EVET" if (model, kosul) in secili else "HAYIR"])

    # --- bayrak csv ---
    with open(a.bayrak, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["oncelik", "id", "kanun", "madde", "probe", "gold",
                    "agirlikli_karsi_pay", "oy_sayisi", "karsi_modeller", "iddia", "kaynak_alinti"])
        for i, (cid, b) in enumerate(sirali, 1):
            m = meta[cid]
            w.writerow([i, cid, m.get("kanun", ""), m.get("madde", ""), m.get("probe", ""),
                        gold[cid], f"{b['pay']:.3f}", b["oy"], b["modeller"],
                        m.get("iddia", ""), m.get("kaynak_alinti", "")])

    # --- onay dosyasi ---
    if os.path.exists(a.onay):
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = load_workbook(a.onay)
        ws = wb["ONAY"]
        c0 = ws.max_column
        basliklar = ["MODEL_UYARISI", "ONCELIK"]
        for i, h in enumerate(basliklar):
            c = ws.cell(1, c0 + 1 + i, h)
            c.font = Font(name="Arial", size=10, bold=True)
            c.fill = PatternFill("solid", fgColor="D9E2F3")
        kirmizi = PatternFill("solid", fgColor="F4CCCC")
        oncelik_no = {cid: i for i, (cid, _) in enumerate(sirali, 1)}
        n = 0
        for r in range(2, ws.max_row + 1):
            cid = str(ws.cell(r, 1).value)
            if cid in bayrak:
                b = bayrak[cid]
                c = ws.cell(r, c0 + 1, f"DIKKAT: agirlikli %{b['pay']*100:.0f} karsi ({b['oy']} bilgili oy)")
                c.font = Font(name="Arial", size=9, color="B22222")
                c.alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(r, c0 + 2, oncelik_no[cid]).font = Font(name="Arial", size=10, bold=True)
                ws.cell(r, 5).fill = kirmizi
                n += 1
        ws.column_dimensions[get_column_letter(c0 + 1)].width = 34
        ws.column_dimensions[get_column_letter(c0 + 2)].width = 9
        ws.auto_filter.ref = f"A1:{get_column_letter(c0 + 2)}{ws.max_row}"

        wt = wb.create_sheet("TANI")
        wt.append(["model", "kosul", "n", "kapsam", "baskin", "baskin_pay", "accD", "accY",
                   "bacc", "J", "elendi", "gerekce", "oy_kullandi"])
        for c in range(1, 14):
            wt.cell(1, c).font = Font(name="Arial", size=10, bold=True)
            wt.cell(1, c).fill = PatternFill("solid", fgColor="D9E2F3")
        for (model, kosul), t in sorted(taniler.items()):
            wt.append([model, kosul, t["n"], round(t["kapsam"], 3), t["baskin"],
                       round(t["baskin_pay"], 3), f(t["acc_D"], 3), f(t["acc_Y"], 3),
                       f(t["bacc"], 3), f(t["J"], 3), "EVET" if t["elendi"] else "HAYIR",
                       t["gerekce"], "EVET" if (model, kosul) in secili else "HAYIR"])
        for c, wdt in enumerate([32, 6, 7, 8, 9, 10, 8, 8, 8, 8, 8, 34, 11], 1):
            wt.column_dimensions[get_column_letter(c)].width = wdt

        os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
        wb.save(a.out)
        print(f"\nyazildi: {a.out} ({n} satir isaretli, ONCELIK sutunu ekli)")
    else:
        print(f"\n! {a.onay} bulunamadi; xlsx adimi atlandi.")
    print(f"yazildi: {a.tani}")
    print(f"yazildi: {a.bayrak}")
    print(f"yazildi: {a.rapor}  <- bu dosyanin icerigini sohbete yapistirin")


if __name__ == "__main__":
    main()
