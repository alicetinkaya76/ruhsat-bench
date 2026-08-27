# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — UZMAN KITABI DOGRULAYICI (uzmanlara gitmeden once).

Neden: kitaplar iki insana gidiyor ve her biri ~1,5 saat harcayacak. Bir
sizinti ya da bicim hatasi ancak dosyalar dondukten sonra fark edilirse
uc saatlik insan emegi cope gider. Bu betik dosyalari YERELDE denetler ve
sohbete yapistirilabilir bir konsol raporu basar; dosyayi gondermek
gerekmez.

Denetimler
----------
  [1] Yapi        : sayfalar, basliklar, satir sayisi, bos hucre
  [2] Kimlik      : iki kitabin kod kumesi ayni mi, anahtarla ortusuyor mu
  [3] KORLUK      : altin etiket / probe / probe_alt / tabaka hicbir
                    hucrede gorunuyor mu (sizinti taramasi)
  [4] SIRA        : iki kitabin sirasi FARKLI mi (yorgunluk etkisinin iki
                    kodlayicida eslesmemesi icin)
  [5] Acilir liste: KARAR ve KALITE sutunlarinda ve her veri satirinda mi
  [6] Tuzak       : anahtardaki tuzak satirlari kitapta var mi
  [7] Turkce      : karakter bozulmasi var mi
  [8] Doluluk     : dagitim oncesi bos olmali

--kendi-testi: sizinti taramasinin POZITIF KONTROLU. Bellekte kasten
altin etiket sizdirilmis bir kopya kurulur; tarama bunu yakalamazsa
tarama arizalidir ve betik bunu bildirir.

Kullanim:
    python scripts/kitap_dogrula.py
    python scripts/kitap_dogrula.py --kendi-testi
"""
import argparse
import csv
import hashlib
import os
import re
import sys

# Iki gecis iki farkli sema kullanir. Betik semayi baslıklardan tanir.
SEMALAR = {
    "gecis1": ["sira", "kod", "belge", "alıntının maddesi", "İDDİA",
               "KAYNAK ALINTISI", "ek bilgi", "KARAR", "KALİTE",
               "GEREKÇE / DÜZELTME", "UZMAN"],
    "gecis2": ["sira", "kod", "belge", "alıntının maddesi", "İDDİA",
               "KAYNAK ALINTISI (başlangıç noktası)", "KARAR",
               "GEREKÇE / kaynakta gördüğünüz", "UZMAN"],
}
# Uzmanin YAZDIGI sutunlar sizinti taramasindan muaftir.
GIRDI_SUTUN = {"KARAR", "KALİTE", "GEREKÇE / DÜZELTME",
               "GEREKÇE / kaynakta gördüğünüz", "UZMAN", "kod", "sira"}
SIZINTI_DESEN = re.compile(r"\bP[1-6]_[a-zçğıöşü_]+|\bA_rastgele\b|\bB_bayrakli\b", re.I)


def oku(yol):
    from openpyxl import load_workbook
    wb = load_workbook(yol)
    ws = wb["DENETIM"]
    bas = [c.value for c in ws[1]]
    satirlar = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        satirlar.append(["" if v is None else str(v) for v in r])
    dv = []
    for d in ws.data_validations.dataValidation:
        dv.append((d.formula1, sum(len(list(rng.cells)) for rng in d.cells.ranges)))
    return wb.sheetnames, bas, satirlar, dv


# Anahtarda bulunabilecek, kitapta GORUNMEMESI gereken alanlar.
GIZLI_ALAN = ("gold", "probe", "probe_alt", "tabaka", "gecis1_kalite")


def sizinti_tara(satirlar, anahtar, bas):
    """Altin/probe/tabaka bilgisi gorunur hucrelerde mi? Bulgulari dondurur."""
    bulgu = []
    kod_i = bas.index("kod") if "kod" in bas else 1
    for s in satirlar:
        kod = s[kod_i]
        a = anahtar.get(kod)
        if not a:
            continue
        gizli = {a[k].strip() for k in GIZLI_ALAN if a.get(k, "").strip()}
        altin = (a.get("gold") or "").strip()
        for j, hucre in enumerate(s):
            if bas[j] in GIRDI_SUTUN:
                continue
            h = hucre.strip()
            if not h:
                continue
            if altin and h == altin:
                bulgu.append((kod, bas[j], "ALTIN ETIKET", h))
            elif h in gizli:
                bulgu.append((kod, bas[j], "PROBE/TABAKA", h))
            if SIZINTI_DESEN.search(h):
                bulgu.append((kod, bas[j], "DESEN", h[:60]))
    return bulgu


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dosyalar",
                    default="data/iddialar/denetim_INS_MUH.xlsx,data/iddialar/denetim_ISG_UZM.xlsx")
    ap.add_argument("--anahtar", default="sonuclar/denetim_anahtar_v2.csv")
    ap.add_argument("--out", default="sonuclar/kitap_dogrulama.txt")
    ap.add_argument("--kendi-testi", action="store_true")
    a = ap.parse_args()

    L = []

    def e(s=""):
        L.append(s)
        print(s)

    hata = 0

    def kontrol(ad, ok, detay=""):
        nonlocal hata
        if not ok:
            hata += 1
        e(f"  [{'GECTI' if ok else 'KALDI'}] {ad}" + (f"   {detay}" if detay else ""))

    with open(a.anahtar, encoding="utf-8-sig") as fh:
        anahtar = {r["kod"]: r for r in csv.DictReader(fh)}

    e("=" * 78)
    e("RUHSAT-Bench — UZMAN KITABI DOGRULAMA")
    e("=" * 78)
    e(f"anahtar: {a.anahtar}  ({len(anahtar)} kod)")

    kitaplar = {}
    for yol in [y.strip() for y in a.dosyalar.split(",") if y.strip()]:
        if not os.path.exists(yol):
            e(f"! dosya yok: {yol}")
            hata += 1
            continue
        kitaplar[yol] = oku(yol)

    if len(kitaplar) < 2:
        e("! iki kitap da okunamadi.")
        return

    for yol, (sayfalar, bas, satirlar, dv) in kitaplar.items():
        ad = os.path.basename(yol)
        e()
        e(f"--- {ad}  ({os.path.getsize(yol)} bayt)")
        kontrol("sayfalar DENETIM + BENIOKU", set(sayfalar) == {"DENETIM", "BENIOKU"}, str(sayfalar))
        sema = next((ad for ad, b in SEMALAR.items() if bas == b), None)
        kontrol("basliklar bilinen bir semayla eslesiyor",
                sema is not None, sema or str(bas))
        kalite_var = "KALİTE" in bas
        kontrol("satir sayisi anahtarla ayni", len(satirlar) == len(anahtar),
                f"{len(satirlar)} / {len(anahtar)}")
        ii, ai = bas.index("İDDİA"), 5
        bos = [s[1] for s in satirlar if not s[ii].strip() or not s[ai].strip()]
        kontrol("bos IDDIA / KAYNAK yok", not bos, f"bos: {bos[:5]}")
        ki = bas.index("KARAR") if "KARAR" in bas else 7
        kontrol("KARAR sutunu bos (dagitim oncesi)",
                all(not s[ki].strip() for s in satirlar),
                f"dolu: {sum(1 for s in satirlar if s[ki].strip())}")
        kar = [n for f, n in dv if "DOGRU" in (f or "")]
        kal = [n for f, n in dv if "TEMIZ" in (f or "")]
        kontrol("KARAR acilir listesi tum satirlarda",
                bool(kar) and kar[0] >= len(satirlar), f"{kar}")
        if kalite_var:
            kontrol("KALITE acilir listesi tum satirlarda",
                    bool(kal) and kal[0] >= len(satirlar), f"{kal}")
        else:
            e("  [ATLANDI] KALITE ekseni bu semada yok")
        tr = "".join(bas)
        kontrol("Turkce karakter bozulmasi yok",
                "İDDİA" in tr and "Ä" not in tr and "Ð" not in tr and "Å" not in tr)

    yollar = list(kitaplar)
    k1, k2 = kitaplar[yollar[0]], kitaplar[yollar[1]]
    kod1 = [s[1] for s in k1[2]]
    kod2 = [s[1] for s in k2[2]]
    e()
    e("--- ORTAK")
    kontrol("iki kitabin kod KUMESI ayni", set(kod1) == set(kod2),
            f"fark: {sorted(set(kod1) ^ set(kod2))[:8]}")
    kontrol("kodlar anahtarla birebir", set(kod1) == set(anahtar),
            f"fark: {sorted(set(kod1) ^ set(anahtar))[:8]}")
    kontrol("SIRA farkli (kor tasarim)", kod1 != kod2,
            f"h1={hashlib.md5(','.join(kod1).encode()).hexdigest()[:8]} "
            f"h2={hashlib.md5(','.join(kod2).encode()).hexdigest()[:8]}")
    tuzak = [k for k, v in anahtar.items()
             if v.get("tuzak") == "1" or v.get("kontrol") == "1"]
    if tuzak:
        kontrol("kontrol/tuzak satirlari kitapta", set(tuzak) <= set(kod1),
                f"{len(tuzak)} kontrol maddesi")
    else:
        e("  [ATLANDI] anahtarda kontrol/tuzak maddesi yok")

    e()
    e("--- KORLUK (sizinti taramasi)")
    toplam = []
    for yol, (_, bas, satirlar, _) in kitaplar.items():
        b = sizinti_tara(satirlar, anahtar, bas)
        toplam += b
        e(f"  {os.path.basename(yol):<28} bulgu: {len(b)}")
        for x in b[:5]:
            e(f"      #{x[0]} | {x[1]} | {x[2]} | {x[3]}")
    kontrol("gorunur hucrelerde altin/probe/tabaka YOK", not toplam, f"{len(toplam)} bulgu")

    if a.kendi_testi:
        e()
        e("--- POZITIF KONTROL (tarama gercekten yakaliyor mu?)")
        _, bas, satirlar, _ = k1
        sahte = [list(s) for s in satirlar]
        kod_i = bas.index("kod")
        # Semaya gore TARANAN bir sutun sec (uzman girdisi olanlar taranmaz).
        hedef_sutun = next(j for j, h in enumerate(bas) if h not in GIRDI_SUTUN)
        hedef = sahte[0][kod_i]
        sahte[0][hedef_sutun] = anahtar[hedef].get("gold", "")
        sahte[1][hedef_sutun] = anahtar[sahte[1][kod_i]].get("probe", "")
        e(f"  sizinti '{bas[hedef_sutun]}' sutununa enjekte edildi")
        b = sizinti_tara(sahte, anahtar, bas)
        kontrol("kasten sizdirilan altin yakalandi",
                any(x[2] == "ALTIN ETIKET" for x in b))
        kontrol("kasten sizdirilan probe yakalandi",
                any(x[2] in ("PROBE/TABAKA", "DESEN") for x in b))

    e()
    e("=" * 78)
    if hata:
        e(f"SONUC: {hata} DENETIM KALDI — kitaplari dagitmayin.")
    else:
        e("SONUC: butun denetimler gecti. Kitaplar dagitilabilir.")
        e(f"  {len(anahtar)} satir, uzman basina ~{len(anahtar)*35//60} dakika.")
    e("=" * 78)

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(L) + "\n")
    print(f"\nyazildi: {a.out}")
    sys.exit(1 if hata else 0)


if __name__ == "__main__":
    main()
