# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — IKINCI GECIS BIRLESTIRME (baglamda dogruluk).

NE OLCULUYOR
------------
Birinci gecis, iddiayi kaynak ALINTISIYLA karsilastirmayi olcuyordu ve
60/60 uyum verdi. Ikinci gecis kaynak MADDENIN TAMAMINA gitmeyi istiyor.
Burada olculen sey, alintinin baglamindan koparildiginda anlamini koruyup
korumadigidir — birinci gecisin yapisal olarak goremedigi sinif.

KRITER AYRISMASI (bu betigin en onemli ciktisi)
-----------------------------------------------
Uyusmazliklar rastgele degilse, iki uzman farkli bir KURAL uyguluyordur ve
bu bir olcum hatasi degil TANIM sorunudur; ortalamayla kapatilamaz, uzlasi
gerektirir. Betik yonu sinar:

  * Uyusmazliklarin cogu tek yonde ise -> sistematik kriter farki.
    Isaret testi (iki yonlu binom) ile raporlanir.
  * Tipik ayrisma sudur:
      A kriteri: "hukum kaynakta gecıyor mu"          -> DOGRU der
      B kriteri: "hukum bu kosullarla mi gecıyor"     -> YANLIS der
    Ikincisi bu gecisin sordugu sorudur; ilki birinci gecisin sorusudur.

POZITIF KONTROL
---------------
K tabakasi, temizlikte atilmis ve altin hatasi DOGRULANMIS maddelerdir.
Yakalama orani bu gecisin GUCUDUR. Dusukse, birinci gecisteki 0/150
sonucu "altin temiz" diye okunamaz; "arac gormuyor" diye okunur.

Kural secimi de buradan cikar: kontrol maddesi uzmanlarin YALNIZCA BIRI
tarafindan yakalandiysa, "iki uzman da altindan farkli" (KESIN) kurali
gercek bir hatayi kacirmis demektir ve birincil olcut GEVSEK kural olmalidir.

Kullanim:
    python scripts/gecis2_birlestir.py --dosyalar data/iddialar/gecis2_INS_MUH_doldurulmus.xlsx,data/iddialar/gecis2_ISG_UZM_doldurulmus.xlsx
"""
import argparse
import csv
import math
import os
from collections import Counter, defaultdict

KESIN = {"DOGRU", "YANLIS"}


def wilson(k, n):
    if n == 0:
        return (float("nan"), float("nan"))
    z, ph = 1.96, k / n
    d = 1 + z * z / n
    m = (ph + z * z / (2 * n)) / d
    r = z * math.sqrt(ph * (1 - ph) / n + z * z / (4 * n * n)) / d
    return (max(0.0, m - r), min(1.0, m + r))


def fisher(a_, b_, c_, d_):
    n = a_ + b_ + c_ + d_
    if n == 0:
        return float("nan")

    def p(x):
        return math.comb(a_ + b_, x) * math.comb(c_ + d_, a_ + c_ - x) / math.comb(n, a_ + c_)
    p0 = p(a_)
    alt, ust = max(0, a_ + c_ - (c_ + d_)), min(a_ + b_, a_ + c_)
    return min(1.0, sum(p(x) for x in range(alt, ust + 1) if p(x) <= p0 + 1e-12))


def isaret_testi(b, c):
    n = b + c
    if n == 0:
        return 1.0
    k = min(b, c)
    return min(1.0, 2 * sum(math.comb(n, i) for i in range(k + 1)) / 2 ** n)


def kappa(x, y):
    n = len(x)
    if n == 0:
        return None
    kats = sorted(set(x) | set(y))
    m = Counter(zip(x, y))
    po = sum(m[(k, k)] for k in kats) / n
    sat = {k: sum(m[(k, b)] for b in kats) / n for k in kats}
    sut = {k: sum(m[(a, k)] for a in kats) / n for k in kats}
    pe = sum(sat[k] * sut[k] for k in kats)
    return None if abs(1 - pe) < 1e-12 else ((po - pe) / (1 - pe), po, pe, n)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dosyalar", required=True)
    ap.add_argument("--anahtar", default="sonuclar/ikinci_gecis_anahtar.csv")
    ap.add_argument("--anahtar1", default="sonuclar/denetim_anahtar_v2.csv",
                    help="birinci gecis anahtari; agirliklarin BILESIKLENMESI icin")
    ap.add_argument("--out", default="sonuclar/gecis2_raporu.txt")
    ap.add_argument("--uzlasi", default="sonuclar/gecis2_uzlasi.csv")
    a = ap.parse_args()

    from openpyxl import load_workbook

    with open(a.anahtar, encoding="utf-8-sig") as fh:
        anahtar = {r["kod"]: r for r in csv.DictReader(fh)}
    # BILESIK AGIRLIK: ikinci gecis, birinci gecis ORNEGINDEN cekildi.
    # Yalnizca ikinci gecis agirligi kullanilirsa tahmin 138 maddelik cerceveye
    # aittir, kiyaslamaya DEGIL. Birinci gecisin tasarim agirligiyla carpilinca
    # tahmin butun kumeye yansitilir.
    agirlik1 = {}
    if os.path.exists(a.anahtar1):
        with open(a.anahtar1, encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                try:
                    agirlik1[r["kod"]] = float(r.get("agirlik") or 1.0)
                except ValueError:
                    agirlik1[r["kod"]] = 1.0

    veri, adlar = {}, []
    for yol in [y.strip() for y in a.dosyalar.split(",") if y.strip()]:
        ws = load_workbook(yol, data_only=True)["DENETIM"]
        b = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}
        ger = next(k for k in b if k.startswith("GEREKÇE"))
        ad = os.path.splitext(os.path.basename(yol))[0].replace("gecis2_", "")
        adlar.append(ad)
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[b["kod"]] is None:
                continue
            veri.setdefault(str(r[b["kod"]]).strip(), {})[ad] = (
                str(r[b["KARAR"]] or "").strip(), str(r[b[ger]] or "").strip())

    L = []

    def e(s=""):
        L.append(s)
        print(s)

    A, B = adlar[0], adlar[-1]
    ortak = [k for k in anahtar if k in veri and all(veri[k].get(x, ("",))[0] for x in adlar)]
    e("=" * 78)
    e("RUHSAT-Bench — IKINCI GECIS BIRLESTIRME")
    e("=" * 78)
    e(f"uzman: {', '.join(adlar)} | anahtarda {len(anahtar)} kod | ikisinde dolu {len(ortak)}")

    # ---------------------------------------------------------------- [1]
    e()
    e("[1] BAGIMSIZLIK")
    ger_ort = [k for k in ortak if veri[k][A][1].strip() and veri[k][B][1].strip()]
    ger_ayni = sum(1 for k in ger_ort if veri[k][A][1].strip() == veri[k][B][1].strip())
    kar_ayni = sum(1 for k in ortak if veri[k][A][0] == veri[k][B][0])
    e(f"    KARAR ayni: {kar_ayni}/{len(ortak)} | GEREKCE metni ayni: {ger_ayni}/{len(ger_ort)}")
    if ger_ort and ger_ayni / len(ger_ort) >= 0.30:
        e("    ! BAGIMSIZ DEGIL — asagisi okunmaz.")
        return

    # ---------------------------------------------------------------- [2]
    e()
    e("[2] POZITIF KONTROL (K tabakasi — bilinen altin hatasi)")
    kontrol = [k for k in ortak if anahtar[k].get("kontrol") == "1"]
    yakalayan = defaultdict(int)
    for k in sorted(kontrol):
        g = anahtar[k]["gold"]
        e(f"    {k}  altin={g}")
        for ad in adlar:
            kr, gr = veri[k][ad]
            tuttu = kr in KESIN and kr != g
            yakalayan[ad] += tuttu
            e(f"      {ad:<16} {kr:<14} {'YAKALADI' if tuttu else 'altinla ayni'}"
              f"   {gr[:70]}")
    e()
    for ad in adlar:
        e(f"    {ad:<16} {yakalayan[ad]}/{len(kontrol)} bilinen hatayi yakaladi")
    ikisi = sum(1 for k in kontrol
                if all(veri[k][ad][0] in KESIN and veri[k][ad][0] != anahtar[k]["gold"]
                       for ad in adlar))
    biri = sum(1 for k in kontrol
               if any(veri[k][ad][0] in KESIN and veri[k][ad][0] != anahtar[k]["gold"]
                      for ad in adlar))
    e(f"    KESIN kural (ikisi birden): {ikisi}/{len(kontrol)}")
    e(f"    GEVSEK kural (en az biri) : {biri}/{len(kontrol)}")
    if kontrol and ikisi < biri:
        e()
        e("    ! KESIN kural bilinen bir hatayi KACIRDI, GEVSEK kural yakaladi.")
        e("      Birincil olcut GEVSEK kural olmalidir; KESIN kural alt sinir verir.")
        e("      Birinci gecisteki 0/150 sonucu da bu isikta okunmalidir.")

    # ---------------------------------------------------------------- [3]
    olcum = [k for k in ortak if anahtar[k].get("kontrol") != "1"]
    e()
    e("[3] KODLAYICILAR ARASI UYUM")
    r = kappa([veri[k][A][0] for k in olcum], [veri[k][B][0] for k in olcum])
    if r:
        k_, po, pe, n = r
        e(f"    Cohen kappa = {k_:.3f}   po={po:.3f} pe={pe:.3f} n={n}")
    ad_ = [k for k in olcum if veri[k][A][0] != veri[k][B][0]]
    yon = Counter((veri[k][A][0], veri[k][B][0]) for k in ad_)
    e(f"    uyusmazlik: {len(ad_)}   yon dagilimi: {dict(yon)}")
    if len(yon) == 2:
        (p1, p2), (q1, q2) = list(yon)
        b_, c_ = yon[(p1, p2)], yon[(q1, q2)]
    else:
        b_, c_ = (list(yon.values()) + [0])[0], 0
    p = isaret_testi(b_, c_)
    e(f"    yon simetrisi (isaret testi) p = {p:.4f}")
    if p < 0.05:
        e()
        e("    ! UYUSMAZLIK SISTEMATIK, rastgele degil. Iki uzman farkli bir")
        e("      KURAL uyguluyor. Bu bir olcum hatasi degil TANIM sorunudur ve")
        e("      ortalamayla kapatilamaz; uzlasi toplantisi gerekir.")
        e("      Tipik ayrisma: 'hukum kaynakta geciyor mu' (DOGRU) ile")
        e("      'hukum BU KOSULLARLA mi geciyor' (YANLIS) arasindadir.")
        e("      Ikincisi bu gecisin sordugu sorudur.")

    # ---------------------------------------------------------------- [4]
    e()
    e("[4] BAGLAMSAL HATA ORANI  (tabaka agirlikli)")
    for kural, ad_kural in (("gevsek", "GEVSEK (en az bir uzman altindan farkli)"),
                            ("kesin", "KESIN  (iki uzman da altindan farkli)")):
        e(f"    {ad_kural}")
        hucre = defaultdict(lambda: [0, 0, 0, 0])   # n, hata, N_tabaka, n_tabaka
        for k in olcum:
            g = anahtar[k]["gold"]
            kr = [veri[k][ad][0] for ad in adlar]
            if kural == "kesin":
                h = len(set(kr)) == 1 and kr[0] in KESIN and kr[0] != g
            else:
                h = any(x in KESIN and x != g for x in kr)
            t = anahtar[k]["tabaka"]
            hucre[t][0] += 1
            hucre[t][1] += h
            hucre[t][2] = int(anahtar[k].get("N_tabaka") or 0)
            hucre[t][3] = int(anahtar[k].get("n_tabaka") or 0)
        Ntop = sum(v[2] for v in hucre.values())
        toplam_olay = sum(v[1] for v in hucre.values())
        toplam_n = sum(v[0] for v in hucre.values())
        pw = var = 0.0
        for t, (n_, x_, N_, _) in sorted(hucre.items()):
            ph = x_ / n_ if n_ else 0.0
            lo, hi = wilson(x_, n_)
            e(f"      {t:<4} {x_:>3}/{n_:<4} = %{100*ph:5.1f}  [%{100*lo:.1f}, %{100*hi:.1f}]"
              f"   cerceve {N_}, agirlik {N_/max(n_,1):.2f}")
            if Ntop:
                pw += (N_ / Ntop) * ph
                if n_ > 1:
                    var += (N_ / Ntop) ** 2 * ph * (1 - ph) / n_ * max(0.0, 1 - n_ / max(N_, 1))
        se = math.sqrt(var)
        if toplam_olay == 0:
            # Sifir olayda normal yaklasim GA'yi noktaya cokertir; YANLISTIR.
            _, ust = wilson(0, toplam_n)
            e(f"      CERCEVE TAHMINI: hicbir olay gozlenmedi (0/{toplam_n})")
            e(f"      %95 TEK YONLU UST SINIR: %{100*ust:.1f}")
            e("      '%0.0 [%0.0, %0.0]' diye raporlamayin.")
        else:
            e(f"      CERCEVE TAHMINI: %{100*pw:.1f}  "
              f"%95GA [%{100*max(0,pw-1.96*se):.1f}, %{100*min(1,pw+1.96*se):.1f}]"
              f"   (birinci gecis ornegi = {Ntop} madde)")
        # bilesik agirlik: butun kumeye yansitma
        if agirlik1 and toplam_olay:
            pay = payda = 0.0
            for k in olcum:
                g = anahtar[k]["gold"]
                kr = [veri[k][ad][0] for ad in adlar]
                h = (len(set(kr)) == 1 and kr[0] in KESIN and kr[0] != g) if kural == "kesin" \
                    else any(x in KESIN and x != g for x in kr)
                t = anahtar[k]["tabaka"]
                n_, _, N_, _ = hucre[t]
                w = (N_ / max(n_, 1)) * agirlik1.get(k, 1.0)
                payda += w
                pay += w * h
            e(f"      BILESIK AGIRLIKLA (butun kumeye yansitilmis): %{100*pay/payda:.1f}")
            e("      Bu, iki asamali ornekleme duzeltmesidir; ustteki cerceve")
            e("      tahmini yalnizca birinci gecis ornegine aittir.")
        e()

    # ---------------------------------------------------------------- [5]
    e("[5] BIRINCI GECIS KALITE BAYRAGININ YORDAYICI GECERLILIGI")
    e("    N tabakasi = birinci geciste kusurlu bulunanlar, T = temiz bulunanlar.")
    gr = defaultdict(lambda: [0, 0])
    for k in olcum:
        g = anahtar[k]["gold"]
        kr = [veri[k][ad][0] for ad in adlar]
        h = any(x in KESIN and x != g for x in kr)
        gr[anahtar[k]["tabaka"]][0] += 1
        gr[anahtar[k]["tabaka"]][1] += h
    nN, xN = gr.get("N", [0, 0])
    nT, xT = gr.get("T", [0, 0])
    if nN and nT:
        pN, pT = xN / nN, xT / nT
        p = fisher(xN, nN - xN, xT, nT - xT)
        e(f"    N: {xN}/{nN} = %{100*pN:.1f}   T: {xT}/{nT} = %{100*pT:.1f}"
          f"   kaldirac x{(pN/pT) if pT else float('inf'):.1f}   Fisher p = {p:.4f}")
        if p < 0.05:
            e("    => KALITE bayragi baglamsal hatayi ONGORUYOR; triyaj icin kullanilabilir.")
        else:
            e("    => Yon dogru ama bu ornek buyuklugunde ANLAMLI DEGIL. Kaldiraci")
            e("       'egilim' olarak raporlayin, yordayici gecerlilik iddia etmeyin.")

    os.makedirs(os.path.dirname(a.uzlasi) or ".", exist_ok=True)
    with open(a.uzlasi, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["kod", "tabaka", "probe", "kanun", "madde", "gold"]
                   + [f"{x}_karar" for x in adlar] + [f"{x}_gerekce" for x in adlar]
                   + ["UZLASI", "KURAL_NOTU"])
        for k in sorted(ad_ + kontrol):
            r_ = anahtar[k]
            w.writerow([k, r_["tabaka"], r_.get("probe", ""), r_.get("kanun", ""),
                        r_.get("madde", ""), r_["gold"]]
                       + [veri[k][x][0] for x in adlar]
                       + [veri[k][x][1] for x in adlar] + ["", ""])
    e()
    e(f"[6] UZLASI DOSYASI: {len(set(ad_ + kontrol))} madde -> {a.uzlasi}")
    e("    KURAL_NOTU sutununa once KURALI yazin (hangi kriter dogru kabul edildi),")
    e("    sonra UZLASI sutununu doldurun. Kural once kararlastirilmazsa uzlasi")
    e("    madde madde pazarlik olur ve tekrarlanabilir olmaz.")

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(L) + "\n")
    print(f"\nyazildi: {a.out}")


if __name__ == "__main__":
    main()
