# -*- coding: utf-8 -*-
"""
RUHSAT-Bench F3 — iki uzman dosyasini birlestirir, Cohen kappa ve
AGIRLIKLI altin hata orani uretir.

Ne hesaplar
-----------
[1] Kapsam: her uzman kac madde doldurdu.
[2] Dikkat tuzagi (varsa): uzmanin gercekten okuyup okumadiginin pozitif
    kontrolu. Yakalama orani dusukse asagidaki her sey zayiflar.
[3] Cohen kappa:
      k3  : {DOGRU, YANLIS, EMIN_DEGILIM} uzerinde
      k2  : yalnizca iki uzmanin da kesin karar verdigi maddelerde
      kQ  : KALITE ekseninde
    Her biri icin asimptotik standart hata ve %95 guven araligi.
[4] Altin gecerliligi:
      KESIN hata  = iki uzman AYNI etiketi verdi ve bu altindan farkli
      ADAY  hata  = en az bir uzman altindan farkli
    Tabakali ornekleme yapildigi icin oranlar TASARIM AGIRLIGIYLA
    (N_hucre/n_hucre) hesaplanir; duz ortalama yanlidir.
    Varyans: tabakali, sonlu evren duzeltmeli.
[5] A vs B kaldiraci: konsensus bayragi gercekten altin hatasi ongoruyor
    mu? Fisher kesin testi. HANDOVER Bolum 8 madde 9'un cevabi budur.
[6] Prob kirilimi ve KALITE dagilimi.
[7] Uyusmazlik listesi -> uzlasi toplantisi icin CSV.

Kullanim:
    python scripts/kappa_birlestir.py --dosyalar data/iddialar/denetim_UZMAN_1.xlsx,data/iddialar/denetim_UZMAN_2.xlsx
"""
import argparse
import csv
import math
import os
from collections import Counter, defaultdict

KESIN = {"DOGRU", "YANLIS"}
GECERLI_KARAR = {"DOGRU", "YANLIS", "EMIN_DEGILIM"}
GECERLI_KALITE = {"TEMIZ", "BAGLAMSIZ", "KIRLI"}
DIAKRITIK = str.maketrans("ĞŞİÇÖÜığşiçöü", "GSICOUIgsicou")


def nrm(s):
    """DOĞRU -> DOGRU. Yalnizca yazim normalizasyonu; olcek cevirisi YAPMAZ."""
    return (s or "").strip().upper().translate(DIAKRITIK).replace(" ", "_")


def kappa(a, b):
    """Cohen kappa + asimptotik SE. a,b esit uzunlukta etiket listeleri."""
    n = len(a)
    if n == 0:
        return None
    kats = sorted(set(a) | set(b))
    idx = {k: i for i, k in enumerate(kats)}
    m = [[0] * len(kats) for _ in kats]
    for x, y in zip(a, b):
        m[idx[x]][idx[y]] += 1
    po = sum(m[i][i] for i in range(len(kats))) / n
    sat = [sum(r) / n for r in m]
    sut = [sum(m[i][j] for i in range(len(kats))) / n for j in range(len(kats))]
    pe = sum(sat[i] * sut[i] for i in range(len(kats)))
    if abs(1 - pe) < 1e-12:
        return (float("nan"), float("nan"), float("nan"), float("nan"), po, pe, n)
    k = (po - pe) / (1 - pe)
    # Fleiss-Cohen-Everitt asimptotik varyans
    t1 = sum(m[i][i] / n * (1 - (sat[i] + sut[i]) * (1 - k)) ** 2 for i in range(len(kats)))
    t2 = (1 - k) ** 2 * sum(m[i][j] / n * (sut[i] + sat[j]) ** 2
                            for i in range(len(kats)) for j in range(len(kats)) if i != j)
    t3 = (k - pe * (1 - k)) ** 2
    var = (t1 + t2 - t3) / (n * (1 - pe) ** 2)
    se = math.sqrt(var) if var > 0 else 0.0
    return (k, se, k - 1.96 * se, k + 1.96 * se, po, pe, n)


def yorum(k):
    if k != k:
        return "tanimsiz"
    for esik, ad in [(0.81, "cok iyi"), (0.61, "iyi"), (0.41, "orta"),
                     (0.21, "zayif"), (0.0, "cok zayif")]:
        if k >= esik:
            return ad
    return "uyum yok"


def fisher(a, b, c, d):
    """2x2 iki yonlu Fisher kesin testi."""
    n = a + b + c + d
    if n == 0:
        return float("nan")
    def p(x):
        return (math.comb(a + b, x) * math.comb(c + d, a + c - x)
                / math.comb(n, a + c))
    alt = max(0, a + c - (c + d))
    ust = min(a + b, a + c)
    p0 = p(a)
    return min(1.0, sum(p(x) for x in range(alt, ust + 1) if p(x) <= p0 + 1e-12))


def wilson(k, n):
    if n == 0:
        return (float("nan"), float("nan"))
    z, ph = 1.96, k / n
    d = 1 + z * z / n
    m = (ph + z * z / (2 * n)) / d
    r = z * math.sqrt(ph * (1 - ph) / n + z * z / (4 * n * n)) / d
    return (max(0.0, m - r), min(1.0, m + r))


def _kapi_testi():
    """[1b] kapisinin kendi kontrolu: bagimliyi engellemeli, bagimsizi GECIRMELI."""
    import random
    rnd = random.Random(7)
    n = 150
    print("=" * 78)
    print("[1b] KAPI TESTI")
    print("=" * 78)
    for ad, ayni_metin, tam_uyum in [
            ("A) kopyalanmis (engellenmeli)", True, True),
            ("B) bagimsiz ama KARAR tam uyumlu (GECMELI)", False, True),
            ("C) bagimsiz, normal ayrisma (GECMELI)", False, False)]:
        ger_ort = list(range(112))
        ger_ayni = len(ger_ort) if ayni_metin else 0
        kar_ayni = n if tam_uyum else n - 9
        ger_oran = ger_ayni / len(ger_ort)
        ger_yeterli = len(ger_ort) >= max(10, int(0.10 * n))
        bagimsiz = (ger_oran < 0.30) if ger_yeterli else not (kar_ayni == n)
        beklenen = (ad[0] != "A")
        print(f"  {ad:<46} bagimsiz={bagimsiz}  "
              f"{'GECTI' if bagimsiz == beklenen else 'KALDI <-- KAPI ARIZALI'}")
    print("\nKapinin dayanagi serbest metindir; tek eksende tam uyum tek basina")
    print("bagimlilik kaniti degildir.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dosyalar", default="", help="virgulle ayrilmis xlsx yollari")
    ap.add_argument("--anahtar", default="sonuclar/denetim_anahtar_v2.csv")
    ap.add_argument("--out", default="sonuclar/kappa_raporu.txt")
    ap.add_argument("--uyusmazlik", default="sonuclar/uyusmazliklar.csv")
    ap.add_argument("--kendi-testi", action="store_true",
                    help="[1b] kapisini sentetik veriyle sinar (pozitif + negatif kontrol)")
    a = ap.parse_args()

    if a.kendi_testi:
        _kapi_testi()
        return
    if not a.dosyalar:
        ap.error("--dosyalar gerekli (ya da --kendi-testi)")

    from openpyxl import load_workbook

    with open(a.anahtar, encoding="utf-8-sig") as fh:
        anahtar = {r["kod"]: r for r in csv.DictReader(fh)}

    kodlar, adlar = {}, []
    for yol in [y.strip() for y in a.dosyalar.split(",") if y.strip()]:
        ws = load_workbook(yol, data_only=True)["DENETIM"]
        bas = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}
        ad = os.path.splitext(os.path.basename(yol))[0].replace("denetim_", "")
        adlar.append(ad)
        for satir in ws.iter_rows(min_row=2, values_only=True):
            kod = str(satir[bas["kod"]]) if satir[bas["kod"]] is not None else None
            if not kod:
                continue
            kr = nrm(satir[bas["KARAR"]]) if satir[bas["KARAR"]] else ""
            kl = nrm(satir[bas["KALİTE"]]) if satir[bas["KALİTE"]] else ""
            gr = satir[bas["GEREKÇE / DÜZELTME"]] if "GEREKÇE / DÜZELTME" in bas else ""
            kodlar.setdefault(kod, {})[ad] = (kr, kl, gr or "")

    L = []

    def e(s=""):
        L.append(s)
        print(s)

    e("=" * 78)
    e("RUHSAT-Bench — UZMAN DENETIMI BIRLESTIRME")
    e("=" * 78)
    e(f"uzman: {', '.join(adlar)} | anahtar: {len(anahtar)} kod")

    # ------------------------------------------------------------ [1]
    e()
    e("[0] GIRDI DOGRULAMA  (asagidaki her sey buna bagli)")
    hard = []
    for ad in adlar:
        var = {k for k, v in kodlar.items() if v.get(ad, ("", "", ""))[0]}
        eksik = set(anahtar) - var
        fazla = var - set(anahtar)
        e(f"    {ad}")
        e(f"      anahtarda {len(anahtar)} kod | dosyada {len(var)} | EKSIK {len(eksik)} | anahtar disi {len(fazla)}")
        if eksik:
            sirali = sorted(eksik, key=lambda x: (not x.isdigit(), int(x) if x.isdigit() else 0))
            e(f"      eksik kodlar: {', '.join(sirali[:20])}")
            tz = [k for k in eksik if anahtar[k].get("tuzak") == "1"]
            if tz:
                e(f"      ! eksiklerin {len(tz)} tanesi DIKKAT TUZAGI: {', '.join(sorted(tz))}")
            hard.append(f"{ad}: {len(eksik)} satir eksik")
        if fazla:
            e(f"      anahtarda olmayan kod: {', '.join(sorted(fazla)[:20])}")
            hard.append(f"{ad}: anahtar disi kod")
        kr = Counter(v[ad][0] for v in kodlar.values() if ad in v and v[ad][0])
        kl = Counter(v[ad][1] for v in kodlar.values() if ad in v and v[ad][1])
        bilinmeyen_kr = {x for x in kr if x not in GECERLI_KARAR}
        bilinmeyen_kl = {x for x in kl if x not in GECERLI_KALITE}
        e(f"      KARAR  degerleri: {dict(kr)}")
        if bilinmeyen_kr:
            e(f"      ! TANINMAYAN KARAR degeri: {sorted(bilinmeyen_kr)}  (beklenen {sorted(GECERLI_KARAR)})")
            hard.append(f"{ad}: KARAR olcegi uyumsuz")
        e(f"      KALITE degerleri: {dict(kl)}")
        if bilinmeyen_kl:
            e(f"      ! TANINMAYAN KALITE degeri: {sorted(bilinmeyen_kl)}  (beklenen {sorted(GECERLI_KALITE)})")
            e("        KALITE bir OLCEK cevirisi degildir; farkli bir olcek kullanildiysa")
            e("        bu eksen tasarlandigi seyi olcmuyordur.")
            hard.append(f"{ad}: KALITE olcegi uyumsuz")
        # eksenler bagimsiz mi
        cift = [(v[ad][0], v[ad][1]) for v in kodlar.values()
                if ad in v and v[ad][0] and v[ad][1]]
        if cift:
            esl = {}
            tek_yonlu = True
            for x, y in cift:
                if esl.setdefault(x, y) != y:
                    tek_yonlu = False
            if tek_yonlu and len(set(x for x, _ in cift)) > 1:
                e(f"      ! KALITE, KARAR'in birebir fonksiyonu: {esl}")
                e("        Iki eksen ayni seyi olcuyor. KALITE ekseni bagimsiz bilgi")
                e("        tasimiyor; yanlis bir iddia PEKALA temiz bir olcum maddesi olabilir.")
                hard.append(f"{ad}: KALITE ekseni KARAR ile ozdes")
    if hard:
        e()
        e("    ! GIRDI DOGRULAMA KALDI:")
        for h in hard:
            e(f"      - {h}")
        e("    Asagidaki istatistikler bu veriyle raporlanabilir DEGILDIR.")
    else:
        e("    => girdi beklenen bicimde.")

    e()
    e("[1] KAPSAM")
    for ad in adlar:
        dolu = sum(1 for v in kodlar.values() if v.get(ad, ("", "", ""))[0])
        dq = sum(1 for v in kodlar.values() if v.get(ad, ("", "", ""))[1])
        e(f"    {ad:<12} KARAR dolu {dolu:>4} / {len(anahtar)}   KALITE dolu {dq:>4}")
    A, B = adlar[0], adlar[-1]
    ortak = [k for k in anahtar
             if all(kodlar.get(k, {}).get(ad, ("", "", ""))[0] for ad in adlar)]
    e(f"    iki uzmanda da dolu: {len(ortak)}")
    if not ortak:
        e("  ! ortak doldurulmus madde yok, cikiliyor.")
        return

    # ------------------------------------------------------------ [1b]
    # BAGIMSIZLIK KONTROLU — kappa'nin on sarti.
    # Cohen kappa iki BAGIMSIZ kodlayici varsayar. Dosyalar birbirinden
    # kopyalanmis ya da tek bir gecisten cogaltilmissa kappa mekanik olarak
    # 1.00 cikar ve HICBIR SEY olcmez. Serbest metin (GEREKCE) bu ayrimi
    # kesin yapar: iki uzmanin ayni gerekceyi kelimesi kelimesine yazma
    # olasiligi pratikte sifirdir.
    e()
    e("[1b] BAGIMSIZLIK KONTROLU  (kappa'nin on sarti)")
    kar_ayni = sum(1 for k in ortak if kodlar[k][A][0] == kodlar[k][B][0])
    kal_ort = [k for k in ortak if kodlar[k][A][1] and kodlar[k][B][1]]
    kal_ayni = sum(1 for k in kal_ort if kodlar[k][A][1] == kodlar[k][B][1])
    ger_ort = [k for k in ortak if kodlar[k][A][2].strip() and kodlar[k][B][2].strip()]
    ger_ayni = sum(1 for k in ger_ort
                   if kodlar[k][A][2].strip() == kodlar[k][B][2].strip())
    e(f"    KARAR birebir ayni : {kar_ayni}/{len(ortak)}")
    if kal_ort:
        e(f"    KALITE birebir ayni: {kal_ayni}/{len(kal_ort)}")
    e(f"    GEREKCE metni birebir ayni: {ger_ayni}/{len(ger_ort)}"
      if ger_ort else "    GEREKCE: iki dosyada birden dolu satir yok")
    ger_oran = (ger_ayni / len(ger_ort)) if ger_ort else 0.0
    # Kanit hiyerarsisi: serbest metin VARSA belirleyici odur. Iki kisinin ayni
    # gerekceyi kelimesi kelimesine yazma olasiligi pratikte sifirdir; buna
    # karsilik TEK BIR EKSENDE %100 uyum, arac karar icin gereken kaniti
    # ekranda gosteriyorsa YAPISAL olarak beklenir ve bagimlilik kaniti DEGILDIR.
    ger_yeterli = len(ger_ort) >= max(10, int(0.10 * len(ortak)))
    if ger_yeterli:
        bagimsiz = ger_oran < 0.30
        e(f"    dayanak: serbest metin ortusmesi %{100*ger_oran:.0f} "
          f"({len(ger_ort)} ortak dolu satirda)")
    else:
        bagimsiz = not (len(ortak) > 20 and kar_ayni == len(ortak)
                        and (not kal_ort or kal_ayni == len(kal_ort)))
        e("    dayanak: serbest metin yetersiz -> yalnizca eksen uyumuna bakildi (ZAYIF)")
    if bagimsiz and kar_ayni == len(ortak) and len(ortak) > 20:
        e()
        e("    (not) KARAR ekseninde uyum TAM. Bagimsizlik serbest metinle")
        e("    dogrulandigi icin bu bir sorun degil, ama kappa=1.00 yorumlanmali:")
        e("    kitap kaynak alintisini ve alintinin maddesini gosteriyorsa cogu")
        e("    prob icin karar bir HATIRLAMA degil KARSILASTIRMA isidir ve tam uyum")
        e("    beklenir. Bu eksen, uretecin manipulasyonunun gosterilen kanittan")
        e("    geri okunabildigini dogrular; altin etiketi mevzuata karsi BAGIMSIZ")
        e("    olarak dogrulamaz. Makalede bu ayrim yazilmali.")
    if not bagimsiz:
        e()
        e("    ! BAGIMSIZ DEGIL. Iki dosya ayni degerlendirmeyi tasiyor.")
        e("      Bu kosullarda Cohen kappa mekanik olarak 1.00 cikar ve")
        e("      kodlayicilar arasi guvenilirlik hakkinda hicbir sey soylemez.")
        e("      Makaleye yazilirsa uydurma bir istatistik olur.")
        e("      KAPPA HESAPLANMADI. Iki BAGIMSIZ doldurma gerekiyor.")
        e()
        e("    Tek gecisten yine de cikarilabilecekler asagida (kappa haric).")
    else:
        e("    => bagimsiz gorunuyor, kappa hesaplanabilir.")

    # ------------------------------------------------------------ [2]
    tuzak = [k for k in ortak if anahtar[k].get("tuzak") == "1"]
    if tuzak:
        e()
        e("[2] DIKKAT TUZAGI (pozitif kontrol — dogru etiket YANLIS)")
        for ad in adlar:
            d = sum(1 for k in tuzak if kodlar[k][ad][0] == "YANLIS")
            e(f"    {ad:<12} {d}/{len(tuzak)} yakalandi")
        e("  Yakalama dusukse asagidaki hata orani tahminleri ALT SINIRDIR.")

    olcum = [k for k in ortak if anahtar[k].get("tuzak") != "1"]

    # ------------------------------------------------------------ [3]
    e()
    e("[3] KODLAYICILAR ARASI UYUM (Cohen kappa)")
    if not bagimsiz:
        e("    ATLANDI — [1b] bagimsizlik kontrolu kaldi.")
    k3 = kappa([kodlar[k][A][0] for k in olcum], [kodlar[k][B][0] for k in olcum])
    ikili = [k for k in olcum if kodlar[k][A][0] in KESIN and kodlar[k][B][0] in KESIN]
    k2 = kappa([kodlar[k][A][0] for k in ikili], [kodlar[k][B][0] for k in ikili])
    kq = [k for k in olcum if kodlar[k][A][1] and kodlar[k][B][1]]
    kQ = kappa([kodlar[k][A][1] for k in kq], [kodlar[k][B][1] for k in kq])
    for ad, r in ([] if not bagimsiz else
                  [("KARAR (3 kategori)", k3), ("KARAR (EMIN_DEGILIM haric)", k2),
                   ("KALITE (3 kategori)", kQ)]):
        if not r:
            e(f"    {ad:<28} hesaplanamadi")
            continue
        k, se, lo, hi, po, pe, n = r
        e(f"    {ad:<28} k={k:5.3f}  %95GA [{lo:5.3f}, {hi:5.3f}]  "
          f"po={po:.3f} pe={pe:.3f}  n={n}   ({yorum(k)})")
    ed = {ad: sum(1 for k in olcum if kodlar[k][ad][0] == "EMIN_DEGILIM") for ad in adlar}
    e(f"    EMIN_DEGILIM sayisi: " + ", ".join(f"{ad}={n}" for ad, n in ed.items()))

    # ------------------------------------------------------------ [4]
    e()
    e("[4] ALTIN GECERLILIGI  (tasarim agirlikli)")

    def hata(k, mod):
        g = anahtar[k]["gold"]
        v = [kodlar[k][ad][0] for ad in adlar]
        if mod == "kesin":
            return len(set(v)) == 1 and v[0] in KESIN and v[0] != g
        return any(x in KESIN and x != g for x in v)

    def agirlikli(kodl, mod):
        hucre = defaultdict(lambda: [0, 0, 0.0, 0])  # n, hata, agirlik, N
        for k in kodl:
            r = anahtar[k]
            h = (r["tabaka"], r["probe"])
            hucre[h][0] += 1
            hucre[h][1] += 1 if hata(k, mod) else 0
            hucre[h][2] = float(r["agirlik"] or 0)
            hucre[h][3] = int(r["N_hucre"] or 0)
        N = sum(v[3] for v in hucre.values())
        if N == 0:
            return (float("nan"), float("nan"), 0, 0)
        p = sum(v[3] * (v[1] / v[0]) for v in hucre.values()) / N
        var = 0.0
        for v in hucre.values():
            n_h, x_h, _, N_h = v
            ph = x_h / n_h
            if n_h > 1:
                fpc = max(0.0, 1 - n_h / N_h)
                var += (N_h / N) ** 2 * ph * (1 - ph) / n_h * fpc
        return (p, math.sqrt(var), sum(v[1] for v in hucre.values()),
                sum(v[0] for v in hucre.values()))

    for mod, ad in [("kesin", "KESIN hata (iki uzman ayni, altindan farkli)"),
                    ("aday", "ADAY  hata (en az bir uzman farkli)")]:
        p, se, x, n = agirlikli(olcum, mod)
        e(f"    {ad}")
        if x == 0:
            # Sifir olayda normal yaklasim GA'yi [0,0]'a cokertir; bu YANLISTIR.
            # Dogru rapor tek yonlu ust sinirdir.
            _, ust = wilson(0, n)
            e(f"      ham: 0/{n}   hicbir hata gozlenmedi")
            e(f"      %95 TEK YONLU UST SINIR: %{100*ust:.2f}")
            e("      NOT: '%0.0 [%0.0, %0.0]' diye raporlamayin. Sifir olay, hata")
            e("      oraninin sifir oldugunu degil, ust sinirin bu degerde oldugunu gosterir.")
        else:
            lo, hi = max(0.0, p - 1.96 * se), min(1.0, p + 1.96 * se)
            e(f"      ham: {x}/{n}   agirlikli oran: %{100*p:.1f}  "
              f"%95GA [%{100*lo:.1f}, %{100*hi:.1f}]")

    # ------------------------------------------------------------ [5]
    e()
    e("[5] A vs B KALDIRACI  (bayrak gercekten hata ongoruyor mu?)")
    for mod in ("kesin", "aday"):
        gr = defaultdict(lambda: [0, 0])
        for k in olcum:
            t = anahtar[k]["tabaka"]
            gr[t][0] += 1
            gr[t][1] += 1 if hata(k, mod) else 0
        nA, xA = gr["A_rastgele"]
        nB, xB = gr["B_bayrakli"]
        pA = xA / nA if nA else float("nan")
        pB = xB / nB if nB else float("nan")
        p = fisher(xB, nB - xB, xA, nA - xA) if nA and nB else float("nan")
        if pA > 0:
            kal = f"x{pB / pA:.2f}"
        elif pB > 0:
            kal = "sonsuz (A'da hata yok)"
        else:
            kal = "iki tabakada da hata yok"
        e(f"    [{mod}] A: {xA}/{nA} = %{100*pA:.1f}   B: {xB}/{nB} = %{100*pB:.1f}"
          f"   kaldirac {kal}   Fisher p={p:.3f}")
        if xA + xB == 0:
            e("      ! IKI TABAKADA DA SIFIR OLAY -> testin gucu SIFIR.")
            e("        Bu sonuc, bayraklarin ise yaramadigini GOSTERMEZ; hicbir sey")
            e("        gostermez. Bayrak konusunda uzman verisiyle cikarim yapilamaz.")
    e("  Olay gozlendiginde: p>0.05, konsensus bayraginin altin hatasi ongormedigi")
    e("  yonunde kanittir. Olay yoksa bu satir okunmaz.")

    # ------------------------------------------------------------ [6]
    e()
    e("[6] PROB KIRILIMI  (A tabakasi, kesin hata)")
    pr = defaultdict(lambda: [0, 0])
    for k in olcum:
        if anahtar[k]["tabaka"] != "A_rastgele":
            continue
        pr[anahtar[k]["probe"]][0] += 1
        pr[anahtar[k]["probe"]][1] += 1 if hata(k, "kesin") else 0
    for p_, (n, x) in sorted(pr.items()):
        lo, hi = wilson(x, n)
        e(f"    {p_:<16} {x:>3}/{n:<3}  %{100*x/n if n else 0:5.1f}  "
          f"[%{100*lo:.1f}, %{100*hi:.1f}]")

    e()
    e("[7] KALITE DAGILIMI")
    for ad in adlar:
        c = Counter(kodlar[k][ad][1] for k in olcum if kodlar[k][ad][1])
        e(f"    {ad:<12} " + "  ".join(f"{x}={n}" for x, n in sorted(c.items())))

    # ------------------------------------------------------------ [8]
    uy = [k for k in olcum if len({kodlar[k][ad][0] for ad in adlar}) > 1]
    os.makedirs(os.path.dirname(a.uyusmazlik) or ".", exist_ok=True)
    with open(a.uyusmazlik, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["kod", "tabaka", "probe", "kanun", "madde", "gold"]
                   + [f"{ad}_karar" for ad in adlar]
                   + [f"{ad}_kalite" for ad in adlar]
                   + [f"{ad}_gerekce" for ad in adlar] + ["UZLASI"])
        for k in sorted(uy):
            r = anahtar[k]
            w.writerow([k, r["tabaka"], r["probe"], r["kanun"], r["madde"], r["gold"]]
                       + [kodlar[k][ad][0] for ad in adlar]
                       + [kodlar[k][ad][1] for ad in adlar]
                       + [kodlar[k][ad][2] for ad in adlar] + [""])
    e()
    e(f"[8] UYUSMAZLIK: {len(uy)} madde -> {a.uyusmazlik}  (UZLASI sutunu birlikte doldurulacak)")

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(L) + "\n")
    print(f"\nyazildi: {a.out}")


if __name__ == "__main__":
    main()
