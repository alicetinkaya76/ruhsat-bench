# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — UZLASI SONRASI NIHAI SAYI (Bolum 3.6).

NEDEN VAR
---------
gecis2_birlestir.py uyusmazliklari `gecis2_uzlasi.csv`'ye yaziyordu ama
DOLU halini geri okuyan bir sey yoktu. Bu betik onu kapatir:

  1. Uzlasi dosyasindaki kararlari alir.
  2. Uyusmazlik OLMAYAN maddelerde iki uzmanin ortak kararini kullanir.
  3. Tabaka agirlikli baglamsal hata oranini YENIDEN hesaplar.
  4. Kontrol maddelerinde (bilinen altin hatasi) yakalama oranini verir.

KAPILAR — asilmadan sayi basilmaz
---------------------------------
  * KURAL_NOTU bos olamaz. Uzlasi kurali yazilmadan madde madde karar
    vermek pazarliktir ve tekrarlanabilir degildir. Betik durur.
  * Uyusmazlik maddelerinin HEPSINDE UZLASI dolu olmalidir.
  * Uzlasi kararlari DOGRU/YANLIS/EMIN_DEGILIM disinda bir sey olamaz.

SIFIR OLAY
----------
Hicbir olay gozlenmezse oran 0 diye raporlanmaz; tek yonlu ust sinir verilir.

Kullanim:
    python scripts/gecis2_uzlasi_isle.py --dosyalar data/iddialar/gecis2_INS_MUH_doldurulmus.xlsx,data/iddialar/gecis2_ISG_UZM_doldurulmus_.xlsx
"""
import argparse
import csv
import math
import os
from collections import defaultdict

KESIN = ("DOGRU", "YANLIS")
GECERLI = ("DOGRU", "YANLIS", "EMIN_DEGILIM")


def wilson(k, n):
    if n == 0:
        return (float("nan"), float("nan"))
    z, ph = 1.96, k / n
    d = 1 + z * z / n
    m = (ph + z * z / (2 * n)) / d
    r = z * math.sqrt(ph * (1 - ph) / n + z * z / (4 * n * n)) / d
    return (max(0.0, m - r), min(1.0, m + r))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--uzlasi", default="sonuclar/gecis2_uzlasi.csv")
    ap.add_argument("--anahtar", default="sonuclar/ikinci_gecis_anahtar.csv")
    ap.add_argument("--anahtar1", default="sonuclar/denetim_anahtar_v2.csv")
    ap.add_argument("--dosyalar", required=True)
    ap.add_argument("--out", default="sonuclar/gecis2_nihai.txt")
    a = ap.parse_args()

    from openpyxl import load_workbook

    L = []

    def e(s=""):
        L.append(s)
        print(s)

    with open(a.anahtar, encoding="utf-8-sig") as fh:
        anahtar = {r["kod"]: r for r in csv.DictReader(fh)}
    agirlik1 = {}
    if os.path.exists(a.anahtar1):
        with open(a.anahtar1, encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                try:
                    agirlik1[r["kod"]] = float(r.get("agirlik") or 1.0)
                except ValueError:
                    agirlik1[r["kod"]] = 1.0

    # uzman kararlari
    uzman = defaultdict(dict)
    for yol in [y.strip() for y in a.dosyalar.split(",") if y.strip()]:
        ws = load_workbook(yol, data_only=True)["DENETIM"]
        b = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}
        ad = os.path.splitext(os.path.basename(yol))[0]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[b["kod"]] is None:
                continue
            uzman[str(r[b["kod"]]).strip()][ad] = str(r[b["KARAR"]] or "").strip()

    with open(a.uzlasi, encoding="utf-8-sig") as fh:
        uz = list(csv.DictReader(fh))

    e("=" * 82)
    e("RUHSAT-Bench — IKINCI GECIS, UZLASI SONRASI NIHAI SAYI")
    e("=" * 82)
    e(f"uzlasi dosyasi: {len(uz)} madde | anahtar: {len(anahtar)} kod")

    # ------------------------------------------------------------- KAPILAR
    e()
    e("[0] KAPILAR")
    kural = [r.get("KURAL_NOTU", "").strip() for r in uz if r.get("KURAL_NOTU", "").strip()]
    if not kural:
        e("    ! KURAL_NOTU HICBIR SATIRDA DOLU DEGIL.")
        e("      Uzlasi kurali yazilmadan sayi hesaplanmaz. Sorulacak soru:")
        e("      'Kaynakta gecen ama kosulu dusurulmus bir iddia DOGRU mu YANLIS mi?'")
        e("      Kural yazilip betik yeniden kosulmalidir.")
        with open(a.out, "w", encoding="utf-8-sig") as fh:
            fh.write("\n".join(L) + "\n")
        return
    e(f"    [GECTI] kural kayitli ({len(kural)} satirda)")
    e(f"      kural: {kural[0][:160]}")

    bos = [r["kod"] for r in uz if not r.get("UZLASI", "").strip()]
    if bos:
        e(f"    ! UZLASI SUTUNU BOS: {len(bos)} madde -> {bos[:12]}")
        with open(a.out, "w", encoding="utf-8-sig") as fh:
            fh.write("\n".join(L) + "\n")
        return
    e(f"    [GECTI] butun uyusmazlik maddelerinde uzlasi dolu ({len(uz)})")

    kotu = [(r["kod"], r["UZLASI"]) for r in uz
            if r["UZLASI"].strip().upper() not in GECERLI]
    if kotu:
        e(f"    ! GECERSIZ UZLASI DEGERI: {kotu}")
        with open(a.out, "w", encoding="utf-8-sig") as fh:
            fh.write("\n".join(L) + "\n")
        return
    e("    [GECTI] butun uzlasi degerleri gecerli")

    # ------------------------------------------------- NIHAI KARAR / MADDE
    uzlasi_karar = {r["kod"]: r["UZLASI"].strip().upper() for r in uz}
    nihai, kaynak = {}, {}
    for kod, kararlar in uzman.items():
        if kod in uzlasi_karar:
            nihai[kod] = uzlasi_karar[kod]
            kaynak[kod] = "uzlasi"
        else:
            v = set(kararlar.values())
            if len(v) == 1:
                nihai[kod] = v.pop()
                kaynak[kod] = "oybirligi"
    e()
    e(f"[1] NIHAI KARAR: {len(nihai)} madde "
      f"({sum(1 for k in kaynak.values() if k=='uzlasi')} uzlasidan, "
      f"{sum(1 for k in kaynak.values() if k=='oybirligi')} oybirliginden)")

    # ------------------------------------------------------------ KONTROL
    e()
    e("[2] KONTROL MADDELERI (bilinen altin hatasi)")
    kontrol = [k for k in nihai if anahtar.get(k, {}).get("kontrol") == "1"]
    yak = 0
    for k in sorted(kontrol):
        g = anahtar[k]["gold"]
        tuttu = nihai[k] in KESIN and nihai[k] != g
        yak += tuttu
        e(f"    {k}  altin={g}  nihai={nihai[k]:<13} "
          f"{'YAKALADI' if tuttu else 'altinla ayni'}  ({kaynak[k]})")
    if kontrol:
        e(f"    uzlasi sonrasi yakalama: {yak}/{len(kontrol)}")

    # -------------------------------------------------------------- ORAN
    olcum = [k for k in nihai if anahtar.get(k, {}).get("kontrol") != "1"]
    e()
    e("[3] BAGLAMSAL HATA ORANI — UZLASI SONRASI")
    hucre = defaultdict(lambda: [0, 0, 0])
    for k in olcum:
        g = anahtar[k]["gold"]
        h = nihai[k] in KESIN and nihai[k] != g
        t = anahtar[k]["tabaka"]
        hucre[t][0] += 1
        hucre[t][1] += h
        hucre[t][2] = int(anahtar[k].get("N_tabaka") or 0)
    Ntop = sum(v[2] for v in hucre.values())
    olay = sum(v[1] for v in hucre.values())
    top_n = sum(v[0] for v in hucre.values())
    pw = var = 0.0
    for t, (n_, x_, N_) in sorted(hucre.items()):
        lo, hi = wilson(x_, n_)
        e(f"    {t:<4} {x_:>3}/{n_:<4} = %{100*x_/n_:5.1f}  "
          f"[%{100*lo:.1f}, %{100*hi:.1f}]   cerceve {N_}")
        if Ntop:
            pw += (N_ / Ntop) * (x_ / n_)
            if n_ > 1:
                var += (N_/Ntop)**2 * (x_/n_)*(1-x_/n_)/n_ * max(0.0, 1 - n_/max(N_, 1))
    se = math.sqrt(var)
    e()
    if olay == 0:
        _, ust = wilson(0, top_n)
        e(f"    hicbir olay gozlenmedi (0/{top_n})")
        e(f"    %95 TEK YONLU UST SINIR: %{100*ust:.1f}")
    else:
        e(f"    CERCEVE TAHMINI: %{100*pw:.1f}  "
          f"%95GA [%{100*max(0,pw-1.96*se):.1f}, %{100*min(1,pw+1.96*se):.1f}]"
          f"   (birinci gecis ornegi = {Ntop} madde)")
        if agirlik1:
            pay = payda = 0.0
            for k in olcum:
                g = anahtar[k]["gold"]
                h = nihai[k] in KESIN and nihai[k] != g
                n_, _, N_ = hucre[anahtar[k]["tabaka"]]
                w = (N_ / max(n_, 1)) * agirlik1.get(k, 1.0)
                payda += w
                pay += w * h
            e(f"    BILESIK AGIRLIKLA (butun kumeye yansitilmis): %{100*pay/payda:.1f}")
            e("    Bolum 3.6'ya bu iki sayi girer: cerceve tahmini araligiyla,")
            e("    bilesik tahmin duyarlilik analizi olarak.")

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(L) + "\n")
    print(f"\nyazildi: {a.out}")


if __name__ == "__main__":
    main()
