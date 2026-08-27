# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — model-konsensüs altın-KG ön-taraması (F3 destek aracı).

Yerel model koşum sonuçlarını (run_local.py çıktısı) üretilen iddia setiyle karşılaştırır;
cevaplayan modellerin ÇOĞUNLUĞUNUN altına karşı çıktığı maddeleri işaretler ve
onay dosyasına MODEL_UYARISI sütunu eklenmiş yeni bir kopya yazar.

Kullanım:
    python scripts/run_local.py --models qwen2.5:7b-instruct-q4_K_M gemma3:4b llama3.2:3b-instruct-q5_K_M gemma3:12b qwen2.5:14b-instruct --claims data/iddialar/uretilen_iddialar_v1.csv --out sonuclar/konsensus.jsonl
    python scripts/konsensus_uyari.py
"""
import argparse, csv, json, os
from collections import defaultdict


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--jsonl", default="sonuclar/konsensus.jsonl")
    ap.add_argument("--csv", default="data/iddialar/uretilen_iddialar_v1.csv")
    ap.add_argument("--onay", default="data/iddialar/onay_turu_v1.xlsx")
    ap.add_argument("--out", default="data/iddialar/onay_turu_v1_uyarili.xlsx")
    ap.add_argument("--esik", type=float, default=0.7, help="karşı-oy oranı eşiği (cevaplayanlar içinde)")
    a = ap.parse_args()

    gold = {}
    for k in csv.DictReader(open(a.csv, encoding="utf-8-sig")):
        gold[k["id"]] = k["gold"]

    oylar = defaultdict(lambda: defaultdict(dict))  # id -> model -> kosul -> karar
    for satir in open(a.jsonl, encoding="utf-8"):
        r = json.loads(satir)
        if r.get("karar") in ("DOGRU", "YANLIS"):
            oylar[r["id"]][r["model"]][r["kosul"]] = r["karar"]

    uyarilar = {}
    for cid, g in gold.items():
        kararlar = []
        for model, kosullar in oylar.get(cid, {}).items():
            # model başına tek oy: E2 varsa E2, yoksa E1
            karar = kosullar.get("E2") or kosullar.get("E1")
            if karar:
                kararlar.append(karar)
        if len(kararlar) < 3:
            continue
        karsi = sum(1 for k in kararlar if k != g)
        oran = karsi / len(kararlar)
        if oran >= a.esik:
            uyarilar[cid] = f"DIKKAT: {karsi}/{len(kararlar)} model altına karşı"
    print(f"{len(gold)} madde | konsensüs uyarısı: {len(uyarilar)} madde (eşik {a.esik:.0%})")
    for cid, u in sorted(uyarilar.items(), key=lambda x: int(x[0]))[:20]:
        print(f"  #{cid}: {u}")
    if len(uyarilar) > 20:
        print(f"  ... (+{len(uyarilar)-20})")

    # Onay dosyasına MODEL_UYARISI sütunu ekle
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    wb = load_workbook(a.onay)
    ws = wb["ONAY"]
    col = ws.max_column + 1
    ws.cell(1, col, "MODEL_UYARISI").font = Font(name="Arial", size=10, bold=True)
    ws.cell(1, col).fill = PatternFill("solid", fgColor="D9E2F3")
    kirmizi = PatternFill("solid", fgColor="F4CCCC")
    n = 0
    for r in range(2, ws.max_row + 1):
        cid = str(ws.cell(r, 1).value)
        if cid in uyarilar:
            c = ws.cell(r, col, uyarilar[cid])
            c.font = Font(name="Arial", size=9, color="B22222")
            ws.cell(r, 5).fill = kirmizi
            n += 1
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col)].width = 26
    wb.save(a.out)
    print(f"yazıldı: {a.out} ({n} satır işaretli) — uzmanlara BU dosyayı verin")


if __name__ == "__main__":
    main()
