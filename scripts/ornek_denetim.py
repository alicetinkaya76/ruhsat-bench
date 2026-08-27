# -*- coding: utf-8 -*-
"""
RUHSAT-Bench F3 — tabakali denetim ornegi ureteci (kor tasarim).

Neden: model-konsensus bayraklari, oy verenler sansa yakin oldugunda altin hatasini
degil probe zorlugunu olcer. Bayrakli satirlari uzmana "oncelik" diye vermek, hem
altin hata oranini yanli tahmin ettirir hem de kiyaslamanin en ayirt edici
maddelerini eleme riskini dogurur. Dogru tasarim: TABAKALI RASTGELE ornek.

Iki tabaka uretir:
  A "rastgele"  : probe'a gore tabakali, tum setten cekilmis rastgele ornek
                  -> altin hata oranini YANSIZ tahmin eder (probe kirilimiyla)
  B "bayrakli"  : yalnizca konsensus bayrakli maddelerden cekilmis ornek
                  -> taramanin gercekten hata ongorup ongormedigini SINAR

Iki tabaka birlestirilip KARISTIRILIR ve uzman dosyasina tabaka bilgisi
YAZILMAZ (kor kodlama). Tabaka esleme anahtari ayri bir dosyaya yazilir;
uzman dosyalari dondukten sonra iki tabakanin hata orani karsilastirilir:
bayrakli tabakada hata orani rastgele tabakadan anlamli olcude yuksek degilse,
tarama ise yaramiyor demektir ve bu makalede raporlanmasi gereken bir bulgudur.

Kullanim:
    python scripts/ornek_denetim.py --mod konsol --n-rastgele 30 --n-bayrak 12
    python scripts/ornek_denetim.py --mod uzman
"""
import argparse
import csv
import json
import os
import random
from collections import defaultdict


def kisalt(s, n):
    s = " ".join((s or "").split())
    return s if len(s) <= n else s[: n - 1] + "…"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default="data/iddialar/uretilen_iddialar_v1.csv")
    ap.add_argument("--bayrak", default="sonuclar/konsensus_bayrak.csv")
    ap.add_argument("--jsonl", default="sonuclar/konsensus.jsonl")
    ap.add_argument("--mod", choices=["konsol", "uzman"], default="konsol")
    ap.add_argument("--n-rastgele", type=int, default=60, help="A tabakasi toplam boyut")
    ap.add_argument("--n-bayrak", type=int, default=30, help="B tabakasi toplam boyut")
    ap.add_argument("--sadece", default="", help="yalniz bu probe (konsol modu icin)")
    ap.add_argument("--kaynak-len", type=int, default=300)
    ap.add_argument("--tohum", type=int, default=20260727)
    ap.add_argument("--out", default="data/iddialar/denetim_ornegi.xlsx")
    ap.add_argument("--anahtar", default="sonuclar/denetim_anahtar.csv")
    ap.add_argument("--konsol-out", default="sonuclar/denetim_konsol.txt")
    a = ap.parse_args()

    with open(a.csv, encoding="utf-8-sig") as fh:
        satirlar = list(csv.DictReader(fh))
    kayit = {k["id"]: k for k in satirlar}

    bayrakli = set()
    if os.path.exists(a.bayrak):
        with open(a.bayrak, encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                bayrakli.add(r["id"])

    oylar = defaultdict(dict)
    if os.path.exists(a.jsonl):
        with open(a.jsonl, encoding="utf-8") as fh:
            for satir in fh:
                satir = satir.strip()
                if not satir:
                    continue
                r = json.loads(satir)
                if r.get("karar"):
                    oylar[str(r["id"])][f"{r['model']}/{r['kosul']}"] = r["karar"][0]

    rnd = random.Random(a.tohum)

    def tabakali(havuz, toplam):
        gruplar = defaultdict(list)
        for cid in havuz:
            gruplar[kayit[cid].get("probe", "?")].append(cid)
        for g in gruplar.values():
            g.sort()
        if not gruplar:
            return []
        pay = max(1, toplam // len(gruplar))
        secim = []
        for p, g in sorted(gruplar.items()):
            secim += rnd.sample(g, min(pay, len(g)))
        kalan = [c for c in sorted(havuz) if c not in set(secim)]
        rnd.shuffle(kalan)
        secim += kalan[: max(0, toplam - len(secim))]
        return sorted(set(secim), key=int)

    A = tabakali([c for c in kayit if c not in bayrakli], a.n_rastgele)
    B = tabakali([c for c in kayit if c in bayrakli], a.n_bayrak) if bayrakli else []
    tabaka = {c: "A_rastgele" for c in A}
    for c in B:
        tabaka[c] = "B_bayrakli"

    os.makedirs(os.path.dirname(a.anahtar) or ".", exist_ok=True)
    with open(a.anahtar, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["id", "tabaka", "probe", "kanun", "gold"])
        for c in sorted(tabaka, key=int):
            k = kayit[c]
            w.writerow([c, tabaka[c], k.get("probe", ""), k.get("kanun", ""), k.get("gold", "")])
    print(f"tabaka A (rastgele): {len(A)} | tabaka B (bayrakli): {len(B)}")
    print(f"anahtar: {a.anahtar}")

    if a.mod == "konsol":
        hedef = [c for c in sorted(tabaka, key=int)
                 if not a.sadece or kayit[c].get("probe") == a.sadece]
        L = []
        for c in hedef:
            k = kayit[c]
            oy = oylar.get(c, {})
            oz = "  ".join(f"{m}={v}" for m, v in sorted(oy.items()))
            L.append(f"--- #{c} | {k.get('kanun','')} | {k.get('madde','')} | "
                     f"{k.get('probe','')} | gold={k.get('gold','')}")
            L.append(f"IDDIA : {kisalt(k.get('iddia',''), 400)}")
            L.append(f"KAYNAK: {kisalt(k.get('kaynak_alinti',''), a.kaynak_len)}")
            if k.get("degisiklik_notu"):
                L.append(f"NOT   : {kisalt(k['degisiklik_notu'], 120)}")
            if oz:
                L.append(f"MODEL : {oz}")
            L.append("")
        metin = "\n".join(L)
        print()
        print(metin)
        os.makedirs(os.path.dirname(a.konsol_out) or ".", exist_ok=True)
        with open(a.konsol_out, "w", encoding="utf-8-sig") as fh:
            fh.write(metin)
        print(f"yazildi: {a.konsol_out}  ({len(hedef)} madde) <- icerigini sohbete yapistirin")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    AR = Font(name="Arial", size=10)
    ARB = Font(name="Arial", size=10, bold=True)
    thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    sirali = sorted(tabaka, key=int)
    rnd.shuffle(sirali)  # kor: tabakalar karisik gelsin

    wb = Workbook()
    ws = wb.active
    ws.title = "DENETIM"
    basliklar = ["sira", "id", "kanun", "madde", "iddia", "kaynak alintisi",
                 "altin etiket", "KARAR", "GEREKCE/DUZELTME", "UZMAN"]
    for c, h in enumerate(basliklar, 1):
        cell = ws.cell(1, c, h)
        cell.font = ARB
        cell.fill = PatternFill("solid", fgColor="D9E2F3")
        cell.border = thin
    dv = DataValidation(type="list", formula1='"ONAY,RET,DUZELT,EMIN_DEGILIM"', allow_blank=True,
                        showDropDown=False, promptTitle="Karar",
                        prompt="ONAY: iddia ve altin etiket dogru | RET: kullanilmasin | DUZELT: kucuk duzeltmeyle kullanilabilir")
    ws.add_data_validation(dv)
    for i, cid in enumerate(sirali, 1):
        k = kayit[cid]
        degerler = [i, cid, k.get("kanun", ""), k.get("madde", ""), k.get("iddia", ""),
                    k.get("kaynak_alinti", ""), k.get("gold", ""), "", "", ""]
        for c, v in enumerate(degerler, 1):
            cell = ws.cell(i + 1, c, v)
            cell.font = AR
            cell.border = thin
            if c in (5, 6):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(i + 1, 8).fill = PatternFill("solid", fgColor="FFF2CC")
        dv.add(ws.cell(i + 1, 8))
    for c, w in enumerate([6, 6, 10, 11, 64, 46, 12, 14, 30, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:J{len(sirali)+1}"

    ws2 = wb.create_sheet("BENIOKU")
    n_probe = len({kayit[c].get("probe") for c in sirali})
    talimat = [
        ("RUHSAT-Bench F3 — Kor Denetim Turu", ""),
        ("", ""),
        ("Gorev", "Her satirda: iddia (E sutunu) kaynak alintisiyla (F) tutarli mi ve altin etiket (G) dogru mu?"),
        ("Onemli", "Satirlarin hangi olcute gore secildigi size bildirilmemistir; bu kasitlidir. Her satiri esit dikkatle degerlendirin."),
        ("ONAY", "Iddia + altin etiket dogru, ifade temiz."),
        ("RET", "Altin etiket yanlis, ifade anlamsiz/kirli ya da baglamsiz."),
        ("DUZELT", "Kucuk mudahaleyle kurtulur; duzeltilmis ifadeyi I sutununa yazin."),
        ("EMIN_DEGILIM", "Ikinci uzmana / karar toplantisina kalsin."),
        ("Cift kodlama", "Bu dosyanin iki kopyasi iki uzmanca BAGIMSIZ doldurulur (J sutununa ad). Cohen kappa raporlanir."),
        ("Kapsam", f"{len(sirali)} satir, {n_probe} farkli probe turu. Sure ~{len(sirali)*20//60} dakika/uzman."),
    ]
    for r, (ad, txt) in enumerate(talimat, 1):
        ws2.cell(r, 1, ad).font = ARB if ad and not txt else AR
        ws2.cell(r, 2, txt).font = AR
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 118

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    wb.save(a.out)
    print(f"yazildi: {a.out}  ({len(sirali)} satir, kor sirali)")
    print("uzmanlara BU dosyayi verin; anahtar dosyasini VERMEYIN.")


if __name__ == "__main__":
    main()
