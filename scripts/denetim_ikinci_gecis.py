# -*- coding: utf-8 -*-
"""
RUHSAT-Bench — IKINCI GECIS: BAGLAMDA DOGRULUK DENETIMI.

NEDEN IKINCI BIR GECIS
----------------------
Birinci gecis 0/150 altin hatasi buldu (%95 ust sinir %2.50). Bu sonuc iki
farkli dunyayla uyumludur ve tek basina ayirt edemez:

  (a) altin kume gercekten cok temiz,
  (b) arac aradigi hatayi gorebilecek turden degil.

(b) ihtimali yapisaldir: birinci gecis kitabi uzmana KAYNAK ALINTISINI
gosteriyor, altin etiket de AYNI alintidan uretiliyor. Uzman kararini
alintiyla karsilastirarak veriyorsa, uretecle birlikte yanilir. Bu tasarimin
goremeyecegi hata sinifi sudur:

    Cumle kendi maddesinden koparildiginda anlamini koruyor mu?

Onceki fikradan tasinan bir kosul, bir anafor ("bu durumda", "sozu edilen"),
ya da maddenin devaminda gelen bir istisna, alintiya bakildiginda gorunmez.

BU GECISIN FARKI
----------------
Soru degisiyor: "iddia alintiyla tutarli mi" degil, "KAYNAK MADDENIN TAMAMINI
ACIP OKUYUN; iddia mevzuata gore hala dogru mu". Uzmandan belgeye gitmesi
istenir. Alinti yine gosterilir ama artik yeterli kanit degil, baslangic
noktasidir.

TABAKALAR
---------
  N — birinci geciste EN AZ BIR uzmanin TEMIZ demedigi maddeler
      (KIRLI veya BAGLAMSIZ). Artik riskin yogunlastigi kume.
  T — birinci geciste IKI uzmanin da TEMIZ dedigi maddelerden rastgele
      ornek. Bu tabaka olmadan yalnizca "bayrakli maddede hata orani"
      olculur; evren tahmini icin ikisi de gerekir.
  K — POZITIF KONTROL: v1'de bulunup temizlikte ATILAN, altin hatasi
      OLDUGU BILINEN maddeler. Uzman bunlari yakalamiyorsa bu gecisin de
      semantik sinifta gucu yok demektir ve 0/150 sonucunun yorumu zayiflar.

Tabaka etiketi uzmana GOSTERILMEZ; siralar karistirilir ve uzman basina
farklidir.

METIN SURUMU
------------
Maddeler ONARILMIS surumden (v6) alinir. Birinci geciste KIRLI isaretlenen
maddelerin bir kismi kelime kirigiydi ve L4 tarafindan duzeltildi; onarilmis
haliyle gostermek, maddenin ESASEN saglam olup olmadigini sinar ve ayni
zamanda onarimi uzman tarafindan dogrular.

Kullanim:
    python scripts/denetim_ikinci_gecis.py
    python scripts/denetim_ikinci_gecis.py --n-temiz 25 --uzman "INS_MUH,ISG_UZM"
"""
import argparse
import csv
import os
import random

# Temizlikte atilan, altin hatasi DOGRULANMIS maddeler (pozitif kontrol).
#   291 — bolum basligi parcasi, onerme degil, yine de gold=DOGRU
#   378 — sayisal manipulasyon capraz atfa denk geldi, olculen sey bozuldu
KONTROL_ID = ["291", "378"]


def oku_kalite(yol):
    from openpyxl import load_workbook
    ws = load_workbook(yol, data_only=True)["DENETIM"]
    b = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value}
    return {str(r[b["kod"]]).strip(): str(r[b["KALİTE"]] or "").strip()
            for r in ws.iter_rows(min_row=2, values_only=True) if r[b["kod"]] is not None}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default="data/iddialar/uretilen_iddialar_v6_onarilmis.csv")
    ap.add_argument("--v1", default="data/iddialar/uretilen_iddialar_v1.csv")
    ap.add_argument("--anahtar1", default="sonuclar/denetim_anahtar_v2.csv")
    ap.add_argument("--dolu", default="data/iddialar/denetim_INS_MUH_DOLDURULMUS.xlsx,"
                                      "data/iddialar/denetim_ISG_UZM_DOLDURULMUS.xlsx")
    ap.add_argument("--n-temiz", type=int, default=25)
    ap.add_argument("--uzman", default="INS_MUH,ISG_UZM")
    ap.add_argument("--tohum", type=int, default=20260729)
    ap.add_argument("--out-dir", default="data/iddialar")
    ap.add_argument("--anahtar", default="sonuclar/ikinci_gecis_anahtar.csv")
    ap.add_argument("--rapor", default="sonuclar/ikinci_gecis_rapor.txt")
    a = ap.parse_args()

    with open(a.csv, encoding="utf-8-sig") as fh:
        kayit = {r["id"]: r for r in csv.DictReader(fh)}
    with open(a.v1, encoding="utf-8-sig") as fh:
        v1 = {r["id"]: r for r in csv.DictReader(fh)}
    with open(a.anahtar1, encoding="utf-8-sig") as fh:
        anahtar1 = {r["kod"]: r for r in csv.DictReader(fh)}

    kaliteler = [oku_kalite(y.strip()) for y in a.dolu.split(",") if y.strip()]
    L = []

    def e(s=""):
        L.append(s)
        print(s)

    e("=" * 78)
    e("RUHSAT-Bench — IKINCI GECIS (baglamda dogruluk)")
    e("=" * 78)
    e(f"metin kaynagi: {a.csv}  (onarilmis surum)")
    e(f"birinci gecis ornegi: {len(anahtar1)} kod")

    gercek = [k for k in anahtar1 if anahtar1[k].get("tuzak") != "1" and k in kayit]
    kirli = [k for k in gercek if any(d.get(k, "") not in ("", "TEMIZ") for d in kaliteler)]
    temiz = [k for k in gercek if k not in kirli]
    rnd = random.Random(a.tohum)
    sec_t = rnd.sample(temiz, min(a.n_temiz, len(temiz)))

    e()
    e("[1] TABAKALAR")
    e(f"    N (en az bir uzman TEMIZ demedi) : {len(kirli):>4} / {len(kirli)} secildi")
    e(f"    T (ikisi de TEMIZ dedi)          : {len(temiz):>4} / {len(sec_t)} secildi")
    kontrol = [i for i in KONTROL_ID if i in v1]
    e(f"    K (bilinen altin hatasi, kontrol): {len(kontrol):>4} / {len(kontrol)} secildi")
    for i in kontrol:
        e(f"        #{i} {v1[i].get('probe','')} gold={v1[i].get('gold','')} "
          f"-> {' '.join((v1[i].get('iddia') or '').split())[:70]}")
    if not kontrol:
        e("    ! KONTROL MADDESI BULUNAMADI. Bu gecisin gucu olculemez;")
        e("      sonuclari ihtiyatla okuyun.")

    kalemler = ([(k, "N", kayit[k]) for k in kirli]
                + [(k, "T", kayit[k]) for k in sec_t]
                + [(f"K{i}", "K", v1[i]) for i in kontrol])
    e(f"  toplam: {len(kalemler)} madde")

    os.makedirs(os.path.dirname(a.anahtar) or ".", exist_ok=True)
    with open(a.anahtar, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["kod", "tabaka", "N_tabaka", "n_tabaka", "agirlik",
                    "probe", "kanun", "madde", "gold", "kontrol",
                    "gecis1_kalite"])
        boyut = {"N": (len(kirli), len(kirli)), "T": (len(temiz), len(sec_t)),
                 "K": (len(kontrol), len(kontrol))}
        for kod, tb, s in kalemler:
            N, n = boyut[tb]
            k1 = "/".join(d.get(kod, "-") for d in kaliteler)
            w.writerow([kod, tb, N, n, f"{N/max(n,1):.4f}", s.get("probe", ""),
                        s.get("kanun", ""), s.get("madde", ""), s.get("gold", ""),
                        1 if tb == "K" else 0, k1 if tb != "K" else "-"])
    e(f"\nanahtar: {a.anahtar}   <-- UZMANLARA VERILMEZ")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    AR, ARB = Font(name="Arial", size=10), Font(name="Arial", size=10, bold=True)
    ince = Border(*[Side(style="thin", color="CCCCCC")] * 4)
    os.makedirs(a.out_dir, exist_ok=True)

    for ui, uzman in enumerate([u.strip() for u in a.uzman.split(",") if u.strip()]):
        sira = list(kalemler)
        random.Random(a.tohum + 977 * (ui + 1)).shuffle(sira)
        wb = Workbook()
        ws = wb.active
        ws.title = "DENETIM"
        bas = ["sira", "kod", "belge", "alıntının maddesi", "İDDİA",
               "KAYNAK ALINTISI (başlangıç noktası)", "KARAR",
               "GEREKÇE / kaynakta gördüğünüz", "UZMAN"]
        for c, h in enumerate(bas, 1):
            cell = ws.cell(1, c, h)
            cell.font, cell.border = ARB, ince
            cell.fill = PatternFill("solid", fgColor="D9E2F3")
        dv = DataValidation(type="list", formula1='"DOGRU,YANLIS,EMIN_DEGILIM"',
                            allow_blank=True, promptTitle="Karar",
                            prompt="Kaynak maddenin TAMAMINI okuduktan sonra karar verin.")
        ws.add_data_validation(dv)
        for i, (kod, tb, s) in enumerate(sira, 1):
            for c, v in enumerate([i, kod, s.get("kanun", ""), s.get("madde", ""),
                                   " ".join((s.get("iddia") or "").split()),
                                   " ".join((s.get("kaynak_alinti") or "").split()),
                                   "", "", uzman], 1):
                cell = ws.cell(i + 1, c, v)
                cell.font, cell.border = AR, ince
                if c in (5, 6):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(i + 1, 7).fill = PatternFill("solid", fgColor="FFF2CC")
            dv.add(ws.cell(i + 1, 7))
        for c, w_ in enumerate([6, 7, 9, 14, 62, 48, 16, 40, 11], 1):
            ws.column_dimensions[get_column_letter(c)].width = w_
        ws.freeze_panes = "E2"

        ws2 = wb.create_sheet("BENIOKU")
        dk = max(1, len(sira) * 60 // 60)
        for r, (ad, txt) in enumerate([
            ("RUHSAT-Bench — İkinci Geçiş", ""),
            ("", ""),
            ("Bu geçiş neden var",
             "Birinci geçişte iddiayı kaynak alıntısıyla karşılaştırdınız. Bu geçişte soru farklı: "
             "alıntı, ait olduğu maddeden koparıldığında anlamını koruyor mu?"),
            ("Ne yapmanız gerekiyor",
             "Her madde için KAYNAK BELGEYİ AÇIN ve alıntının alındığı maddenin TAMAMINI okuyun. "
             "F sütunundaki alıntı yeterli kanıt değil, başlangıç noktasıdır."),
            ("", ""),
            ("KARAR", "Maddenin tamamını okuduktan sonra: iddia mevzuata göre doğru mu?"),
            ("  DOGRU", "İddia, maddenin bütünü içinde de doğrudur."),
            ("  YANLIS", "İddia yanlıştır — ya da alıntı bağlamından koparıldığı için "
                         "maddede olmayan bir anlam kazanmıştır."),
            ("  EMIN_DEGILIM", "Kaynağa baktıktan sonra da karara varamıyorsanız."),
            ("", ""),
            ("Özellikle bakın",
             "Önceki fıkradan taşınan bir koşul; 'bu durumda', 'söz konusu', 'anılan' gibi "
             "bağlam gerektiren ifadeler; maddenin devamında gelen bir istisna; "
             "yalnızca belirli yapı sınıfları veya işyerleri için geçerli olan bir hüküm."),
            ("Gerekçe", "H sütununa kaynakta ne gördüğünüzü yazın — özellikle YANLIS dediyseniz."),
            ("", ""),
            ("Süre", f"{len(sira)} madde × ~1 dk ≈ {dk} dakika."),
            ("Bağımsızlık", "Tek başınıza doldurun; diğer uzmanla dosyalar dolana kadar konuşmayın."),
        ], 1):
            ws2.cell(r, 1, ad).font = ARB if ad and not txt else AR
            ws2.cell(r, 2, txt).font = AR
            ws2.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 108

        yol = os.path.join(a.out_dir, f"gecis2_{uzman}.xlsx")
        wb.save(yol)
        e(f"yazildi: {yol}  ({len(sira)} madde)")

    os.makedirs(os.path.dirname(a.rapor) or ".", exist_ok=True)
    with open(a.rapor, "w", encoding="utf-8-sig") as fh:
        fh.write("\n".join(L) + "\n")
    print(f"yazildi: {a.rapor}")


if __name__ == "__main__":
    main()
